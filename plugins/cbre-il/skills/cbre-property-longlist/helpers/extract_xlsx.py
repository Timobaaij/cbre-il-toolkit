#!/usr/bin/env python3
"""extract_xlsx.py - read an Excel input in one of two modes.

  questionnaire : a client requirements form -> meta.requirements + slider default
                  + hard-requirement flags (NOT property records)
  tracker       : one row per property/scenario -> candidate records (columns
                  alias-mapped to canonical fields)

Mode is auto-detected: a sheet whose header row contains >=3 known property
column names is a tracker; otherwise it is treated as a questionnaire.

The tracker map is WIDE (a real 75-column CRE building tracker mapped only ~7
fields once, so a complete dataset looked thin and everything downstream
degraded). Design rules learned from that failure:
  * collisions resolve by alias priority + header coverage, never first-column
    ("Status" beats "Construction status"; the losers stay as per-row FALLBACK
    columns for empty primary cells). Landlord and developer are DISTINCT fields,
    NOT collision rivals: a "Landlord"/"Owner"/"Asset manager" column maps to its
    own `landlord` field and never backfills `developer` (that conflation shipped a
    landlord as the developer);
  * units live in the HEADER and are honoured - '(sq ft)' areas stay sq ft
    (source convention is KEPT), '£ per sq ft' rents ship as £/sq ft/yr with
    the band for that convention, '(m)' suffixes a bare numeric ('15 m');
  * combined 'Latitude, Longitude' columns split into exact lat/lng (first-party
    coordinates beat any geocoder);
  * a sheet mapping >=8 fields is a RICH TRACKER (__meta.tracker_rich) - merge
    gives its records spec authority over brochures;
  * a parsed AREA is checked against a CONSERVATIVE plausibility band (the twin of
    the rent band: sq m 300-600,000; sq ft 3,000-6,500,000 - deliberately wide, a
    coarse backstop for a 10x unit error or a parse-garble, never policing real
    estate) AND a sq-ft-vs-sq-m magnitude cross-check (a 'sq m' value > 60,000 or a
    'sq ft' value < 4,000 is almost certainly the other unit). Both FLAG-never-convert:
    an out-of-band / suspect area is KEPT (a real figure is never dropped or coerced to
    tbd) with a low-confidence prov note + a header_report flag (area_out_of_band /
    area_unit_suspect) -> the yield/Gaps pipeline for the broker to confirm. NO
    clear-height band (real warehouse clear heights legitimately exceed 24 m);
  * every sheet emits a header_report (populated vs mapped columns + the
    unmapped header names) so a thin parse of a strong source is LOUD.

The column->field DECISION can come from an isolated LLM tracker-interpretation
sub-agent (the same pattern as the brochure interpretation sub-agent): pass a
cached map via --colmap / the column_map= parameter and the LLM names each column
(plus per-column basis / currency / per-area / monthly hints), while THIS module
still PARSES every number with the arithmetic above (the LLM never reads a cell).
The COLUMN_MAP dictionary stays load-bearing as the OFFLINE FALLBACK (column_map
is None), a HARD NEGATIVE-table VETO on the LLM map, a backfill of any column the
LLM left unbound, and a logged cross-check. column_map=None is byte-identical to
the historical dictionary path. See reference/interpretation.md 'Tracker mode'.

CLI:
  python extract_xlsx.py <file.xlsx> [--region R --country C] [--out out.json]
                         [--colmap work/extract/<slug>_<hash8>_map.json]
"""
from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path

from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parent))
try:
    from rapidfuzz import fuzz, process
except Exception:  # sandbox without rapidfuzz: difflib-backed shim
    from rapidfuzz_shim import fuzz, process
import normalize as N
import coords as _CO

# Alias order IS priority: the first alias is the canonical column name, later
# ones are acceptable stand-ins (used per row only when the better column is
# empty). All matching is whole-word, so 'park' never claims 'car parking'.
COLUMN_MAP = {
    "park": ["park", "marketing name", "project", "scheme", "property", "site name",
             "building name", "logistics park", "unit name"],
    "developer": ["developer", "promoter"],
    # LANDLORD is a DISTINCT party from the developer (the owner / asset manager /
    # freeholder of an existing building, vs the party that built/is building it). One
    # field absorbs landlord/owner/asset-manager/freeholder; a separate "owner" is
    # deferred. It NEVER backfills developer (that conflation shipped a landlord as the
    # developer and then split one property into two cards across the difference).
    "landlord": ["landlord", "owner", "asset manager", "freeholder"],
    "city": ["city", "town", "location", "municipality"],
    # country: multilingual but DELIBERATELY excludes bare "land" (DE/NL for country
    # but also = plot/ground -> would mis-bind plotArea); unmapped country falls to
    # the geocoder. English first so UK/IE byte-stability holds.
    "country": ["country", "pais", "país", "pays", "paese", "kraj"],
    "region": ["region", "county", "province"],
    "warehouseArea": ["warehouse area", "gla", "gia", "warehouse sqm", "size",
                      "area sqm", "total area", "floor area", "building size", "total size"],
    "plotArea": ["site area", "plot area", "land area", "plot size", "site size"],
    "officeArea": ["office area", "office content", "office sqm", "offices"],
    "warehouseRentVal": ["rent", "headline rent", "quoting rent", "current quoting rent",
                         "asking rent", "rent eur", "warehouse rent"],
    "serviceCharge": ["service charge"],
    "landPrice": ["land price", "asking price", "guide price", "plot price", "land value"],
    "leaseTerm": ["lease term", "lease terms"],
    "incentives": ["incentive", "incentives"],
    "status": ["status", "availability", "construction status"],
    "earlyAccess": ["early access", "availability date", "delivery", "pc of construction",
                    "practical completion", "completion date"],
    "clearHeight": ["clear height", "eaves", "haunch", "clear internal height", "height"],
    "floorLoad": ["floor loading", "floor load", "slab loading"],
    "loadingDocks": ["dock level doors", "dock doors", "loading docks", "docks"],
    "overheadDoors": ["level access doors", "overhead doors", "drive in doors",
                      "ground level doors", "level access"],
    "electricity": ["power", "power supply", "electricity", "kva"],
    "truckParking": ["trailer spaces", "truck parking", "hgv parking", "lorry parking"],
    "carParking": ["car parking", "car parking spaces", "parking spaces", "car spaces"],
    "breeam": ["breeam", "breeam rating", "certification", "epc rating", "epc"],
    "motorway": ["motorway", "highway", "corridor", "road corridor", "m25 segment"],
    "latlng": ["latitude, longitude", "lat, long", "lat/long", "coordinates",
               "lat lng", "latlong"],
    "lat": ["latitude", "lat"],
    "lng": ["longitude", "lng", "long", "lon"],
}

# headers an alias must NEVER claim, checked before any matching: substring
# matching once bound "Rent free (months)" to warehouseRentVal, shipping a "6"
# (months) as a €6/m²/yr headline rent that no plausibility band can catch.
# 'historical/achieved/per annum/review/deal' keep dated or lump-sum rent
# columns away from the HEADLINE quoting rent; 'unit/ratio' keep 'Size Unit'
# and 'Door ratio' style derived columns out of the data fields.
NEGATIVE = {
    "warehouseRentVal": re.compile(
        r"rent\s*free|free\s*rent|service\s*charge|incentive|deposit"
        r"|historic|achieved|per\s+annum|p\.?a\.?\b|review|deal", re.I),
    "city": re.compile(r"note|comment|remark|description", re.I),
    "warehouseArea": re.compile(r"plot|land|office|unit\b|ratio", re.I),
    "plotArea": re.compile(r"ratio", re.I),
    "clearHeight": re.compile(r"or above|above\?|ratio", re.I),
    "loadingDocks": re.compile(r"ratio|total", re.I),
    "overheadDoors": re.compile(r"ratio|total", re.I),
    "electricity": re.compile(r"charg|ev\b|solar", re.I),
    "status": re.compile(r"verified", re.I),
    "developer": re.compile(r"verified", re.I),
    "landlord": re.compile(r"verified", re.I),
    "leaseTerm": re.compile(r"start|expiry|break|outside", re.I),
    "park": re.compile(r"solus|/", re.I),
    # land PRICE must never claim a rent column (rent / per sq ft / psf / service)
    "landPrice": re.compile(r"rent|per\s*sq|psf|/\s*sq|service", re.I),
}

# a rent column quoted per month is annualised x12 (the conversion is recorded
# in the row's provenance, never hidden)
MONTHLY_HDR = re.compile(r"month|/\s*mo\b|/mes\b|p\.?\s?m\b|monat|mensual|mensuel", re.I)

# A separate 'Size Unit' / 'Area basis' column QUALIFIES the size figure: when its value
# (or the size header itself) says GIA/GEA/GLA/gross, the 'Size' is a GROSS TOTAL, NOT the
# warehouse area - so warehouse = total - office (a real tracker quotes Size=GIA + a
# separate Office content column). Only then; a plain size column defaults to warehouse
# area (the default is applied ONLY AFTER this check). _UNIT_COL_RX finds the qualifier
# column; _GIA_RX matches the gross-total marker in its value or the size header.
_UNIT_COL_RX = re.compile(r"\b(?:size|area|gia|gea|gla|floor(?:\s*area)?|measurement)\s*(?:unit|basis|type)\b", re.I)
_GIA_RX = re.compile(r"\b(?:gia|gea|gla|gross(?:\s+(?:internal|external))?(?:\s+area)?)\b", re.I)

# fields where a tracker's literal 0 is an empty cell in disguise (bulk exports
# zero-fill unknowns; a genuine 0 eaves/power/term does not exist)
_ZERO_IS_UNKNOWN = {"clearHeight", "floorLoad", "loadingDocks", "overheadDoors",
                    "electricity", "truckParking", "carParking", "leaseTerm",
                    "officeArea", "serviceCharge", "incentives", "warehouseArea",
                    "plotArea"}

_LL_SPLIT = re.compile(r"(-?\d{1,2}\.\d{3,})\s*[,;/ ]\s*(-?\d{1,3}\.\d{3,})")

# the skill's OWN Source-Ledger column signature - a sheet carrying all of these is a
# prior deliverable, not a client tracker (P2-1 defensive guard; see ledger.py columns)
_LEDGER_SIG = {"property_id", "record_type", "field", "value", "source_file"}


def _merge_ledger_aliases() -> None:
    """Append label_ledger.json multilingual aliases to the English COLUMN_MAP so
    continental (DE/PL/FR/IT/ES/NL/PT) trackers map - the commonest European input
    used to yield NO dashboard. English stays FIRST and matching is whole-word +
    coverage-ranked, so UK/IE/CEE-English mapping is byte-identical. A few bare tokens
    are BLOCKED from a field where they would mis-bind (the ledger's bare 'Warehouse'
    is a rent label there but must NEVER be the xlsx rent column)."""
    LEDGER_TO_XLSX = {
        "city": "city", "developer": "developer", "landlord": "landlord", "status": "status",
        "earlyAccess": "earlyAccess", "plotArea": "plotArea", "warehouseArea": "warehouseArea",
        "officeArea": "officeArea", "clearHeight": "clearHeight", "floorLoad": "floorLoad",
        "loadingDocks": "loadingDocks", "overheadDoors": "overheadDoors",
        "truckParking": "truckParking", "carParking": "carParking",
        "warehouseRent": "warehouseRentVal", "serviceCharge": "serviceCharge",
        "landPrice": "landPrice", "breeam": "breeam",
    }
    BLOCK = {"warehouseRentVal": {"warehouse"}}
    EXTRA = {"warehouseArea": ["surface"]}  # FR bare 'Surface'; the ledger only has compounds
    try:
        led = json.loads((Path(__file__).resolve().parent.parent / "assets"
                          / "label_ledger.json").read_text(encoding="utf-8")).get("fields", {})
    except Exception:
        led = {}  # ledger absent/corrupt -> English-only mapping, never a crash
    for lf, payload in led.items():
        xf = LEDGER_TO_XLSX.get(lf)
        if not xf or xf not in COLUMN_MAP:
            continue
        block = BLOCK.get(xf, set())
        have = {a.lower() for a in COLUMN_MAP[xf]}
        for a in payload.get("aliases", []):
            al = " ".join(str(a).strip().lower().split())
            if al and al not in have and al not in block:
                COLUMN_MAP[xf].append(al)
                have.add(al)
    for xf, extras in EXTRA.items():
        have = {a.lower() for a in COLUMN_MAP[xf]}
        for al in extras:
            if al not in have:
                COLUMN_MAP[xf].append(al)


_merge_ledger_aliases()


def _norm(s) -> str:
    return " ".join(str(s or "").strip().lower().split())


def _header_candidates(header) -> list[tuple[str, tuple]]:
    """All (field, score) candidates for one header. score = (tier, coverage,
    -alias_pos): tier 3 exact, 2 whole-word, 1 fuzzy>=90; coverage = how much of
    the header the alias explains ('Eaves (m)' beats 'Eaves 10m or above?')."""
    h = _norm(header)
    if not h:
        return []
    out = []
    for field, aliases in COLUMN_MAP.items():
        neg = NEGATIVE.get(field)
        if neg and neg.search(h):
            continue
        best = None
        for pos, a in enumerate(aliases):
            if h == a:
                cand = (3, 1.0, -pos)
            elif re.search(rf"(?<![a-z0-9]){re.escape(a)}(?![a-z0-9])", h):
                cand = (2, len(a) / len(h), -pos)
            elif fuzz.ratio(a, h) >= 90:
                cand = (1, len(a) / len(h), -pos)
            else:
                continue
            if best is None or cand > best:
                best = cand
        if best:
            out.append((field, best))
    return out


def _header_field(header) -> str | None:
    """Best single field for a header (compat surface used by the evals)."""
    cands = _header_candidates(header)
    return max(cands, key=lambda c: c[1])[0] if cands else None


def _map_header_row(row) -> dict[str, list[tuple[int, str, str]]]:
    """field -> ranked [(column index, normalised header, original header), ...].
    One column maps to at most one field (its best candidate); one field may keep
    SEVERAL columns, best first - the later ones serve as per-row fallbacks
    ('Promoter' fills developer when the 'Developer' cell is empty - a WITHIN-field
    fallback; 'Landlord' is a SEPARATE field and never backfills developer)."""
    field_cols: dict[str, list] = {}
    for cidx, cell in enumerate(row):
        if cell in (None, ""):
            continue
        cands = _header_candidates(cell)
        if not cands:
            continue
        field, score = max(cands, key=lambda c: c[1])
        field_cols.setdefault(field, []).append((score, -cidx, cidx, _norm(cell), str(cell)))
    mapping = {}
    for field, cands in field_cols.items():
        cands.sort(reverse=True)
        mapping[field] = [(c[2], c[3], c[4]) for c in cands]
    return mapping


# --------------------------------------------------------------------------- #
# LLM TRACKER MAPPING (the column->field DECISION moves to an isolated sub-agent;
# the dictionary above stays as the OFFLINE FALLBACK, a HARD NEGATIVE-table VETO,
# and a logged cross-check). The LLM returns a MAP ONLY - it never reads a cell
# value; Python parses every number with the EXISTING arithmetic, so every numeric
# guarantee is byte-preserved. See reference/interpretation.md 'Tracker mode'.

# the canonical field names the parse loop emits (the only fields an LLM map may
# bind a column to); 'role' columns (e.g. size_basis) are not fields.
_CANON_FIELDS = set(COLUMN_MAP.keys()) | {"lat", "lng"}


def _hint_header(hdr_orig: str, col: dict) -> str:
    """Augment a column's ORIGINAL header with the LLM's explicit unit/basis hints so
    the UNCHANGED parse arithmetic (N.area_unit_of / N.currency_of / MONTHLY_HDR /
    _GIA_RX) reads them. The hint OVERRIDES the header-regex read only when the LLM is
    explicit; an absent hint leaves the original header tokens (regex fallback). The
    tokens are appended in the source's own convention - currency/per-area/period are
    NEVER converted, only NAMED, so Python applies x12 / GIA-office faithfully."""
    extra = []
    cur = str(col.get("currency") or "").strip().upper()
    if cur in ("GBP", "£"):
        extra.append("£")
    elif cur in ("EUR", "€"):
        extra.append("€")
    per = str(col.get("perArea") or col.get("areaUnit") or "").strip().lower()
    if per in ("sq ft", "sqft", "sq.ft", "ft2", "ft²"):
        extra.append("sq ft")
    elif per in ("sq m", "sqm", "sq.m", "m2", "m²"):
        extra.append("sq m")
    elif per == "acres":
        extra.append("acres")
    elif per == "ha":
        extra.append("ha")
    period = str(col.get("period") or "").strip().lower()
    if period in ("month", "monthly"):
        extra.append("per month")
    basis = str(col.get("basis") or "").strip().lower()
    if basis in ("gia", "gea", "gla", "gross"):
        extra.append(basis if basis != "gross" else "gross")
    return f"{hdr_orig} ({' '.join(extra)})" if extra else str(hdr_orig)


def _llm_best_map(header_row, column_map: dict) -> tuple[dict, dict, int | None]:
    """Build the parse-loop `best_map` (field -> [(col_index, normalised_header,
    original_header), ...]) from an LLM column map, returning (best_map, report, the
    size-basis column index). Each numeric/unit hint is folded into the column's
    header via _hint_header so the EXISTING arithmetic honours it.

    Three dictionary safeguards are enforced here (the dictionary is never deleted):
      * HARD VETO - a column whose ORIGINAL header trips the NEGATIVE regex for the
        LLM-named field can NEVER bind to it; that one binding is dropped and the
        dictionary's own best field for the column is used instead (fall back per
        column, never ship the catastrophic 'months as rent' class).
      * CROSS-CHECK - when the dictionary ALSO confidently maps a column (tier 3/2)
        to a DIFFERENT field, the disagreement is logged for the reviewer.
      * the column->field decision is otherwise the LLM's.
    """
    field_cols: dict[str, list] = {}
    vetoes: list[str] = []
    disagreements: list[str] = []
    size_basis_col = None
    columns = (column_map or {}).get("columns", []) if isinstance(column_map, dict) else []
    for col in columns:
        if not isinstance(col, dict):
            continue
        try:
            cidx = int(col.get("index"))
        except (TypeError, ValueError):
            continue
        if cidx < 0 or cidx >= len(header_row):
            continue
        hdr_orig = "" if header_row[cidx] in (None, "") else str(header_row[cidx])
        role = str(col.get("role") or "").strip().lower()
        if role == "size_basis" and size_basis_col is None:
            size_basis_col = cidx
        field = col.get("field")
        if not field or field not in _CANON_FIELDS:
            continue  # null / non-property / unknown field -> Python falls back per column
        hn = _norm(hdr_orig)
        # HARD VETO: the dictionary NEGATIVE table beats the LLM, always
        neg = NEGATIVE.get(field)
        if neg and hn and neg.search(hn):
            vetoes.append(f"col {cidx} '{hdr_orig}' -> {field} (NEGATIVE-vetoed)")
            continue
        # CROSS-CHECK: a confident dictionary field that disagrees is logged
        dict_field = _header_field(hdr_orig)
        if dict_field and dict_field != field:
            disagreements.append(
                f"col {cidx} '{hdr_orig}': LLM->{field}; dictionary->{dict_field}")
        aug = _hint_header(hdr_orig, col)
        # rank within a field by the LLM's column ORDER (first listed wins, like the
        # dictionary's alias priority), so a later duplicate is a per-row fallback
        field_cols.setdefault(field, []).append((cidx, _norm(aug), aug))
    # VETO FALLBACK + UNMAPPED FALLBACK: for any column the LLM left unbound (null /
    # vetoed / not listed), let the dictionary claim it - so the offline dictionary
    # strength is never LOST by an LLM that maps thin. The dictionary mapping is built
    # from the ORIGINAL header row, and an LLM binding for a column always WINS.
    llm_bound = {ci for cols in field_cols.values() for ci, _h, _o in cols}
    dict_map = _map_header_row(header_row)
    for field, cols in dict_map.items():
        for ci, hd, ho in cols:
            if ci in llm_bound:
                continue
            field_cols.setdefault(field, []).append((ci, hd, ho))
            llm_bound.add(ci)
    best_map = {f: [(c[0], c[1], c[2]) for c in cols] for f, cols in field_cols.items()}
    report = {"map_source": "llm", "llm_dictionary_disagreements": disagreements,
              "llm_vetoes": vetoes}
    return best_map, report, size_basis_col


def _zeroish(raw) -> bool:
    n = N.normalize_number(raw)
    return n is not None and n == 0


def _sheets(path: Path):
    """Yield (title, rows) per sheet. Reads .csv via the stdlib too (common export
    format), and a corrupt/encrypted workbook yields NOTHING rather than raising a
    traceback (run.py then classifies + surfaces the file honestly)."""
    if path.suffix.lower() == ".csv":
        import csv
        try:
            with open(path, newline="", encoding="utf-8-sig") as fh:
                sample = fh.read(8192)
                fh.seek(0)
                try:
                    dialect = csv.Sniffer().sniff(sample) if sample.strip() else csv.excel
                except Exception:
                    dialect = csv.excel
                rows = [tuple(r) for r in csv.reader(fh, dialect)]
            if rows:
                yield (path.stem, rows)
        except Exception:
            return
        return
    try:
        wb = load_workbook(path, data_only=True, read_only=True)
    except Exception:
        return  # corrupt / encrypted / unsupported - no sheets, no traceback
    try:
        for ws in wb.worksheets:
            yield (ws.title, list(ws.iter_rows(values_only=True)))
    finally:
        wb.close()


_LINK_OK = re.compile(r"\.(?:pdf|pptx)(?:[?#]|$)", re.I)


def _harvest_links(path: Path) -> list[dict]:
    """Brochure URLs / hyperlink targets sitting in cells (P2-3). The sandbox cannot
    GET them, but they must not silently vanish - surfaced in the Gaps Report so the
    orchestrator (or the broker) can fetch them. Returns [{locator, target}] sorted
    deterministically. Reads cell VALUES (URL-as-string) and hyperlink OBJECTS."""
    found: dict[str, str] = {}

    def _add(loc, tgt):
        t = str(tgt or "").strip()
        if t and (_LINK_OK.search(t) or t.lower().startswith(("http://", "https://"))):
            found.setdefault(t, loc)
    try:
        if path.suffix.lower() == ".csv":
            for title, rows in _sheets(path):
                for ri, row in enumerate(rows, 1):
                    for ci, c in enumerate(row, 1):
                        if isinstance(c, str):
                            _add(f"{title}!r{ri}c{ci}", c)
        else:
            ro = path.stat().st_size > 8_000_000  # huge workbook: values-only (skip hyperlink objects)
            wb = load_workbook(path, data_only=True, read_only=ro)
            try:
                for ws in wb.worksheets:
                    for row in ws.iter_rows():
                        for cell in row:
                            if isinstance(cell.value, str):
                                _add(f"{ws.title}!{cell.coordinate}", cell.value)
                            hl = getattr(cell, "hyperlink", None)
                            if hl is not None and getattr(hl, "target", None):
                                _add(f"{ws.title}!{cell.coordinate}", hl.target)
            finally:
                wb.close()
    except Exception:
        pass  # link harvest is best-effort; never break extraction
    return [{"locator": loc, "target": tgt}
            for tgt, loc in sorted(found.items(), key=lambda kv: kv[1])]


def _sheet_map(column_map, ws_title: str):
    """Select the LLM column map that applies to one sheet. `column_map` may be a
    bare {columns:[...]} (applied to the file's tracker sheet) or a per-sheet dict
    {sheet_title: {columns:[...]}}. None / no match -> None (dictionary fallback)."""
    if not isinstance(column_map, dict):
        return None
    if "columns" in column_map:
        return column_map
    sub = column_map.get(ws_title)
    return sub if isinstance(sub, dict) and "columns" in sub else None


def detect_and_extract(path: Path, region: str = "", country: str = "",
                       column_map: dict | None = None,
                       column_map_verify: dict | None = None) -> dict:
    """Read an Excel/CSV input. `column_map` (OPTIONAL, default None) is the cached
    LLM tracker map; when None the EXISTING dictionary path runs VERBATIM (byte-
    identical output - the whole offline battery exercises this). When provided for a
    tracker sheet the LLM decides the column->field mapping, the NEGATIVE table still
    vetoes, the dictionary still backfills unbound columns, and Python parses every
    number with the SAME arithmetic (no value ever comes from the model).

    `column_map_verify` (OPTIONAL, default None) is a SECOND, blind, independent
    re-derivation of the SAME map (the semantic verifier - reference/interpretation.md
    "Verification pass"). When present it NEVER drives the parse; it is only diffed
    against `column_map` per sheet, and any per-column field/basis disagreement is
    attached as `hr_entry['semantic_disagreements']` (ADVISORY - surfaced to the broker
    via the yield/Gaps pipeline, the primary map is never rejected). Absent -> no diff,
    byte-identical to today (the offline-fallback invariant)."""
    path = Path(path)
    records, requirements, header_report = [], {}, []
    for ws_title, rows in _sheets(path):
        if not rows:
            continue
        # DEFENSIVE (P2-1): a prior-run Source Ledger left in the inputs folder is NOT
        # a tracker - its rows once became phantom 'requirements'. Refuse any sheet
        # whose header carries the ledger's own column signature (catches a RENAMED
        # ledger that the intake filename filter would miss).
        _hdr_cells = {_norm(c) for row in rows[:3] for c in (row or ()) if c}
        if _LEDGER_SIG.issubset(_hdr_cells):
            continue
        # find the header row with the most known columns in the first 8 rows
        best_hdr, best_map = None, {}
        for ridx in range(min(8, len(rows))):
            mapping = _map_header_row(rows[ridx])
            if len(mapping) > len(best_map):
                best_hdr, best_map = ridx, mapping

        if best_hdr is not None and len(best_map) >= 3:  # TRACKER
            # the dictionary finds the header ROW robustly; an LLM map (when supplied for
            # this sheet) then REPLACES the column->field decision for that same row. The
            # NEGATIVE veto + a dictionary backfill of unbound columns keep the dictionary
            # load-bearing; column_map=None leaves the dictionary path byte-identical.
            llm_map = _sheet_map(column_map, ws_title)
            llm_report = None
            llm_size_basis = None
            if llm_map is not None:
                best_map, llm_report, llm_size_basis = _llm_best_map(rows[best_hdr], llm_map)
            # SEMANTIC VERIFIER (advisory): when a second, blind re-derivation of the map
            # is supplied, diff it against the primary map for THIS sheet by column index.
            # The primary map still drives the parse above; this only surfaces a field/basis
            # disagreement for the broker (never rejects the primary map). Off the critical
            # path - absent verify map -> no diff -> byte-identical to today.
            verify_disagreements = None
            verify_map = _sheet_map(column_map_verify, ws_title)
            if llm_map is not None and verify_map is not None:
                verify_disagreements = diff_tracker_maps(llm_map, verify_map, rows[best_hdr])
            tracker_rich = len(best_map) >= 8
            # the size-qualifier column ('Size Unit'/'Area basis'), if any: its per-row
            # value decides whether the size is a GROSS total (GIA) or the warehouse area.
            # An LLM role:'size_basis' column takes precedence over the header regex.
            gia_unit_col = llm_size_basis if llm_size_basis is not None else next(
                (ci for ci, cell in enumerate(rows[best_hdr])
                 if _UNIT_COL_RX.search(_norm(cell))), None)
            rent_unit_silent = False  # any row shipped with an ASSUMED (defaulted) rent unit
            area_out_of_band = []     # (park, value, unit) of any area outside its plausibility band
            area_unit_suspect = []    # (park, value, unit) of any sq-ft-vs-sq-m magnitude smell
            for roff, r in enumerate(rows[best_hdr + 1:]):
                rownum = best_hdr + 2 + roff  # 1-based sheet row (real locator)
                rec = {"region": region, "country": country}
                prov = {}
                for field, cols in best_map.items():
                    cidx, hdr, hdr_orig = next(
                        ((ci, hd, ho) for ci, hd, ho in cols
                         if ci < len(r) and r[ci] not in (None, "")),
                        (None, None, None))
                    if cidx is None:
                        continue
                    raw = r[cidx]
                    if field in _ZERO_IS_UNKNOWN and _zeroish(raw):
                        continue  # bulk exports zero-fill unknowns - honest absence
                    loc = f"{ws_title}!r{rownum}"
                    if field == "latlng":
                        m = _LL_SPLIT.search(str(raw))
                        if m:
                            lat, lng = float(m.group(1)), float(m.group(2))
                            if -90 <= lat <= 90 and -180 <= lng <= 180 \
                                    and (abs(lat) > 0.01 or abs(lng) > 0.01):
                                rec["lat"], rec["lng"] = lat, lng
                                prov["lat"] = prov["lng"] = f"{loc} (split from '{hdr}')"
                        continue
                    if field in ("lat", "lng"):
                        num = N.normalize_number(raw)
                        lo, hi = (-90, 90) if field == "lat" else (-180, 180)
                        if num is not None and lo <= num <= hi and num != 0:
                            rec[field] = num
                            prov[field] = loc
                        continue
                    if field in ("warehouseArea", "plotArea"):
                        num = N.normalize_number(raw)
                        if num is None or num <= 0:
                            continue
                        unit = N.area_unit_of(hdr)
                        if unit == "acres":  # UK convention pairs acres with sq ft
                            num = round(num * N.SQFT_PER_ACRE)
                            loc += " (acres x43,560 -> sq ft)"
                            unit = "sq ft"
                        elif unit == "ha":
                            num = round(num * N.SQM_PER_HA)
                            loc += " (ha x10,000 -> sq m)"
                            unit = "sq m"
                        rec[field] = num
                        if unit in ("sq ft", "sq m"):
                            rec.setdefault("areaUnit", unit)
                        prov[field] = loc
                        # AREA PLAUSIBILITY + sq-ft-vs-sq-m MAGNITUDE (the area twin of the
                        # rent band/unit smell, on the FINAL stored magnitude post acres/ha).
                        # Unlike the rent band (which DROPS at ~575), an out-of-band/suspect
                        # area is KEPT and SURFACED: a real figure must never be silently
                        # dropped or coerced to tbd, and the magnitude check NEVER auto-converts.
                        resolved_unit = unit if unit in ("sq ft", "sq m") else rec.get("areaUnit")
                        lo, hi = N.area_band_for(resolved_unit)
                        if not (lo <= num <= hi):
                            prov[field] += (f" (AREA OUT OF BAND {lo:g}-{hi:g} "
                                            f"{resolved_unit or 'sq m'}; kept for broker review)")
                            area_out_of_band.append((rec.get("park"), num, resolved_unit or "sq m"))
                        note = N.area_magnitude_mismatch(num, resolved_unit)
                        if note:
                            prov[field] += f" ({note})"
                            area_unit_suspect.append((rec.get("park"), num, resolved_unit or "sq m"))
                        continue
                    if field == "warehouseRentVal":
                        num = N.normalize_number(raw)
                        if num is None:
                            continue
                        cur_hdr, per_hdr = N.currency_of(hdr), N.area_unit_of(hdr)
                        per = per_hdr if per_hdr in ("sq ft", "sq m") else None
                        # P1-3: HEADER unit-silent -> read currency/area from the CELL
                        # ('£8.50 / sq ft'); the header ALWAYS keeps precedence
                        cell_stated = False
                        if cur_hdr is None or per is None:
                            cu = N.rent_unit_of_text(raw)
                            if cu:
                                cell_stated = True
                                parts = cu.split("/")
                                cur_hdr = cur_hdr or (parts[0] or None)
                                if per is None and len(parts) > 1 and parts[1] in ("sq ft", "sq m"):
                                    per = parts[1]
                        unit = N.rent_unit_str(cur_hdr, per)
                        # P1-4: a monthly marker in the HEADER OR the CELL annualises x12
                        if MONTHLY_HDR.search(hdr) or N.MONTHLY_RX.search(str(raw)):
                            num = round(num * 12, 2)
                            loc += f" ({N.normalize_number(raw):g}/mo x12 -> annual)"
                        lo, hi = N.rent_unit_band(unit)
                        if not (lo <= num <= hi):
                            continue  # zero-filled / implausible quoting rent
                        # HONESTY (unit-silent rent): when NEITHER the header NOR the cell
                        # states a currency or per-area, the displayed €/sq m/yr is a house
                        # DEFAULT, not the source. A bare '8.5' under 'Rent' could be a UK
                        # £8.5/sq ft quote - shipping it as €8.5/sq m/yr invents both the
                        # currency (FX = invention) and the unit. We still ship the number
                        # (degrading to tbd would lose a real figure) but mark rentUnit as
                        # ASSUMED in provenance AND raise it on the sheet's header_report so
                        # the yield/Gaps pipeline surfaces it for the broker to confirm. The
                        # value is otherwise byte-identical to today, so any source that DOES
                        # state a unit (header or cell) is completely unaffected.
                        unit_assumed = cur_hdr is None and per is None and not cell_stated
                        rec["warehouseRentVal"] = num
                        rec["rentUnit"] = unit
                        rec["warehouseRent"] = N.rent_display(num, unit)
                        prov[field] = loc
                        prov["warehouseRent"] = loc
                        if unit_assumed:
                            rec["rentUnitAssumed"] = True
                            prov["rentUnit"] = (
                                f"{loc} (rent unit ASSUMED {unit} - source stated no "
                                f"currency or per-area unit; confirm with the landlord/agent)")
                            rent_unit_silent = True
                        continue
                    # string fields; a unit stated in the header suffixes a bare
                    # numeric cell ('Eaves (m)' 15 -> '15 m'); date cells ship ISO
                    import datetime as _dt
                    if isinstance(raw, _dt.datetime):
                        val = raw.date().isoformat()
                    elif isinstance(raw, _dt.date):
                        val = raw.isoformat()
                    else:
                        val = N.clean_value(raw)
                        suffix = N.header_value_suffix(hdr_orig)
                        if suffix and re.fullmatch(r"-?\d[\d .,]*", val):
                            val = f"{val}{suffix}"
                    rec[field] = val
                    prov[field] = loc
                # GIA / gross-total size: when the size that fed warehouseArea is a GROSS
                # total (the 'Size Unit'/'Area basis' column says GIA/GEA/GLA, or the size
                # header itself does), it is NOT the warehouse area - warehouse = GIA -
                # office. Subtract only when an office figure is present; a plain size with
                # no gross marker stays warehouse area (the default, applied only here).
                wh = rec.get("warehouseArea")
                if isinstance(wh, (int, float)):
                    wh_hdr = next((hd for ci, hd, ho in best_map.get("warehouseArea", [])
                                   if ci < len(r) and r[ci] not in (None, "")), "")
                    unit_val = (str(r[gia_unit_col]) if (gia_unit_col is not None
                                and gia_unit_col < len(r) and r[gia_unit_col] not in (None, "")) else "")
                    if _GIA_RX.search(wh_hdr) or _GIA_RX.search(unit_val):
                        office_num = (N.extract_first_number(str(rec.get("officeArea")))
                                      if not N.looks_unknown(rec.get("officeArea")) else None)
                        base = prov.get("warehouseArea", f"{ws_title}!r{rownum}")
                        if office_num and 0 < office_num < wh:
                            rec["warehouseArea"] = round(wh - office_num)
                            prov["warehouseArea"] = (f"{base} (warehouse = GIA {wh:g} - office "
                                                     f"{office_num:g}; size unit GIA)")
                        else:
                            prov["warehouseArea"] = f"{base} (GIA gross total; office not separately stated)"
                # skip footer/aggregate rows ("Total", "Sum", "Average") and rows
                # too thin to be a property - they became phantom records
                park_l = str(rec.get("park", "")).strip().lower()
                if re.fullmatch(r"(?:grand\s+)?total|sum|average|avg\.?", park_l or "-"):
                    continue
                if len(prov) < 2:
                    continue
                if rec.get("park") or rec.get("warehouseArea"):
                    rec["__meta"] = {"source_file": path.name, "source_type": "xlsx",
                                     "locator_base": ws_title, "prov": prov,
                                     "tracker_rich": tracker_rich}
                    # First-party map-link pins: if the row did NOT already yield coordinates from a
                    # dedicated lat/lng column, stash any maps URL / bare 'lat,lng' found in the row's
                    # cells for the shared resolver (extract_pdf.backfill_link_coords) to parse. This
                    # never sets coords here (the resolver owns the parse + precedence) and is bounded
                    # to matching strings, so a large free-text cell is not copied wholesale.
                    if not (isinstance(rec.get("lat"), (int, float))
                            and isinstance(rec.get("lng"), (int, float))):
                        cands = [c.strip() for c in r if isinstance(c, str)
                                 and (_CO.MAPS_URI.search(c) or _CO.PLAIN_LL.search(c))]
                        if cands:
                            rec["__meta"]["map_candidates"] = cands
                    records.append(rec)
            # extraction-yield report: populated columns vs mapped columns. A
            # field-rich sheet yielding a thin parse must be LOUD, not silent.
            data_rows = rows[best_hdr + 1:]
            populated = [(ci, str(c)) for ci, c in enumerate(rows[best_hdr])
                         if c not in (None, "") and any(
                             ci < len(dr) and dr[ci] not in (None, "") for dr in data_rows)]
            mapped_cols = {ci for cols in best_map.values() for ci, _h, _o in cols}
            hr_entry = {
                "sheet": ws_title,
                "populated_columns": len(populated),
                "mapped_columns": sum(1 for ci, _ in populated if ci in mapped_cols),
                "tracker_rich": tracker_rich,
                "unmapped_headers": [h for ci, h in populated if ci not in mapped_cols],
            }
            if rent_unit_silent:  # a rent shipped with a DEFAULTED unit (no currency/unit
                # stated in header or cell) - surfaced so the yield/Gaps pipeline asks the
                # broker to confirm, never a silent €/sq m/yr invention
                hr_entry["rent_unit_assumed"] = True
            if area_out_of_band:  # an area outside its plausibility band - KEPT + surfaced
                # (never dropped/coerced); the yield/Gaps pipeline asks the broker to confirm
                hr_entry["area_out_of_band"] = [
                    {"park": p, "value": v, "unit": u} for p, v, u in area_out_of_band]
            if area_unit_suspect:  # a sq-ft-vs-sq-m magnitude smell - value KEPT, NOT converted
                hr_entry["area_unit_suspect"] = [
                    {"park": p, "value": v, "unit": u} for p, v, u in area_unit_suspect]
            if llm_report is not None:
                # the LLM mapped this sheet: record the source + any dictionary
                # disagreements / NEGATIVE vetoes for the reviewer (a thinner LLM parse than
                # the dictionary still trips the same mapped<populated yield note above)
                hr_entry.update(llm_report)
            if verify_disagreements:  # two independent map passes disagreed on a field/basis
                # ADVISORY - surfaced to the broker via run.py yield_notes -> yield_report.md
                # -> the Gaps Report; the PRIMARY map (above) still drove the parse
                hr_entry["semantic_disagreements"] = verify_disagreements
            header_report.append(hr_entry)
        else:  # fewer than 3 mapped columns
            hdr_row = rows[best_hdr] if best_hdr is not None else (rows[0] if rows else ())
            data_rows = rows[(best_hdr + 1) if best_hdr is not None else 1:]
            populated = [(ci, str(c)) for ci, c in enumerate(hdr_row)
                         if c not in (None, "") and any(
                             ci < len(dr) and dr[ci] not in (None, "") for dr in data_rows)]
            mapped_cols = {ci for cols in (best_map or {}).values() for ci, _h, _o in cols}
            if len(populated) >= 4 and len(data_rows) >= 2:
                # P1-2: a WIDE sheet whose headers we could not map is a TRACKER, not a
                # questionnaire - do NOT fold its rows into requirements (phantom
                # entries); surface the unmapped headers so run.py/the broker sees why
                header_report.append({
                    "sheet": ws_title, "populated_columns": len(populated),
                    "mapped_columns": len(mapped_cols), "tracker_rich": False,
                    "suspected_tracker": True,
                    "unmapped_headers": [h for ci, h in populated if ci not in mapped_cols],
                })
            else:  # genuine narrow questionnaire (<=3 populated columns)
                for r in rows:
                    cells = [c for c in r if c not in (None, "")]
                    if len(cells) >= 2 and isinstance(cells[0], str) and len(str(cells[0])) < 60:
                        requirements[_norm(cells[0])] = N.clean_value(cells[1])
    return {"records": records, "requirements": requirements,
            "header_report": header_report, "linked_sources": _harvest_links(path)}


def _cellstr(c) -> str:
    """Deterministic string form of a cell for the manifest + structure hash: a date
    ships ISO, a number ships :g (no float jitter), everything else is stripped text."""
    import datetime as _dt
    if c in (None, ""):
        return ""
    if isinstance(c, _dt.datetime):
        return c.date().isoformat()
    if isinstance(c, _dt.date):
        return c.isoformat()
    if isinstance(c, float):
        return f"{c:g}"
    return str(c).strip()


def tracker_structure(path: Path, sample_rows: int = 5) -> list[dict]:
    """The per-tracker-sheet STRUCTURE an isolated mapping sub-agent needs, and the
    bytes the cache hash is computed over: [{sheet, headers (raw, in column order),
    sample_rows (up to N data rows, each a list of cell strings), populated_columns,
    unmapped_headers}]. Pure detection (the dictionary finds the header row + its own
    miss list to focus the model); reads NO column_map. A non-tracker sheet
    (questionnaire / ledger / too-thin) is omitted. Deterministic + offline."""
    out: list[dict] = []
    for ws_title, rows in _sheets(Path(path)):
        if not rows:
            continue
        _hdr_cells = {_norm(c) for row in rows[:3] for c in (row or ()) if c}
        if _LEDGER_SIG.issubset(_hdr_cells):
            continue
        best_hdr, best_map = None, {}
        for ridx in range(min(8, len(rows))):
            mapping = _map_header_row(rows[ridx])
            if len(mapping) > len(best_map):
                best_hdr, best_map = ridx, mapping
        if best_hdr is None or len(best_map) < 3:
            continue  # questionnaire / un-detectable - not a tracker sheet
        header_row = rows[best_hdr]
        headers = [_cellstr(c) for c in header_row]
        data_rows = rows[best_hdr + 1:]
        samples = [[_cellstr(c) for c in r] for r in data_rows[:sample_rows]]
        populated = [(ci, str(c)) for ci, c in enumerate(header_row)
                     if c not in (None, "") and any(
                         ci < len(dr) and dr[ci] not in (None, "") for dr in data_rows)]
        mapped_cols = {ci for cols in best_map.values() for ci, _h, _o in cols}
        out.append({
            "sheet": ws_title,
            "headers": headers,
            "sample_rows": samples,
            "populated_columns": len(populated),
            "unmapped_headers": [h for ci, h in populated if ci not in mapped_cols],
        })
    return out


def _load_colmap(p):
    """Load a cached LLM tracker map for --colmap. Accepts the cache wrapper
    {input_hash, schema_version, map:{columns:[...]}} or a bare map; best-effort
    (None on any error, so a corrupt cache degrades to the dictionary, never crashes)."""
    if not p:
        return None
    try:
        obj = json.loads(Path(p).read_text(encoding="utf-8"))
    except Exception:
        return None
    if isinstance(obj, dict) and isinstance(obj.get("map"), dict):
        return obj["map"]
    return obj if isinstance(obj, dict) else None


# the column keys two independent maps are compared on. `field` is the load-bearing
# column->field DECISION; `basis`/`areaUnit`/`currency`/`perArea`/`period` are the
# size/rent BASIS the verifier exists to re-derive (a 'sq m' vs 'sq ft' read of the
# same column, a GIA vs warehouse basis). Order is fixed so the output is byte-stable.
_VERIFY_KEYS = ("field", "basis", "areaUnit", "currency", "perArea", "period")


def _verify_norm(v):
    """Compare two map values case-insensitively, with null/'' folded to None, so a
    cosmetic 'GIA' vs 'gia' or a present-vs-absent key is judged honestly (an absent
    key on ONE side IS a disagreement worth surfacing - the two passes did not agree
    on the basis)."""
    if v is None:
        return None
    s = str(v).strip().lower()
    return s or None


def diff_tracker_maps(map1: dict | None, map2: dict | None,
                      header_row=None) -> list[dict]:
    """Compare a PRIMARY tracker map (the author's) against a VERIFY map (a second,
    blind re-derivation) by COLUMN INDEX and return the list of per-column/key
    disagreements - the deterministic Python DIFF the independent semantic verifier
    rides on. PURE FUNCTION: no I/O, no clock, no randomness; the output is SORTED by
    (index, key) so identical inputs always yield identical bytes.

    Each disagreement is {index, header, key, pass1, pass2} where key is one of
    _VERIFY_KEYS and pass1/pass2 are the two passes' values (null/absent INCLUDED as a
    value - a column one pass mapped and the other left unmapped IS a disagreement).
    The diff is ADVISORY ONLY: it never changes which map drives the parse (the primary
    map still does) and never rejects either map - it surfaces to the broker via the
    yield/Gaps pipeline. Order-independence note: swapping map1/map2 yields the SAME
    disagreement set with pass1/pass2 labels swapped (the comparator is symmetric)."""
    def _by_index(m):
        out = {}
        cols = (m or {}).get("columns", []) if isinstance(m, dict) else []
        for col in cols:
            if not isinstance(col, dict):
                continue
            try:
                idx = int(col.get("index"))
            except (TypeError, ValueError):
                continue
            out[idx] = col  # last entry for an index wins (mirrors a malformed dup)
        return out
    a = _by_index(map1)
    b = _by_index(map2)
    hdr = list(header_row) if header_row is not None else []
    diffs: list[dict] = []
    for idx in set(a) | set(b):
        ca, cb = a.get(idx, {}), b.get(idx, {})
        for key in _VERIFY_KEYS:
            va, vb = ca.get(key), cb.get(key)
            if _verify_norm(va) != _verify_norm(vb):
                header = ""
                if 0 <= idx < len(hdr) and hdr[idx] not in (None, ""):
                    header = str(hdr[idx])
                diffs.append({"index": idx, "header": header, "key": key,
                              "pass1": va, "pass2": vb})
    diffs.sort(key=lambda d: (d["index"], d["key"]))
    return diffs


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx")
    ap.add_argument("--region", default="")
    ap.add_argument("--country", default="")
    ap.add_argument("--out")
    ap.add_argument("--colmap", help="cached LLM tracker column map (work/extract/<slug>_<hash8>_map.json)")
    ap.add_argument("--colmap-verify", dest="colmap_verify",
                    help="a SECOND, blind tracker map (work/extract/<slug>_<hash8>_mapcheck.json) - "
                         "diffed against --colmap for the semantic verifier; never drives the parse")
    args = ap.parse_args()
    res = detect_and_extract(Path(args.xlsx), args.region, args.country,
                             column_map=_load_colmap(getattr(args, "colmap", None)),
                             column_map_verify=_load_colmap(getattr(args, "colmap_verify", None)))
    sys.stdout.reconfigure(encoding="utf-8")
    payload = json.dumps(res, ensure_ascii=False, indent=2)
    if args.out:
        Path(args.out).write_text(payload, encoding="utf-8")
        print(f"OK {len(res['records'])} records, {len(res['requirements'])} requirements -> {args.out}")
        for hr in res.get("header_report", []):
            if hr["mapped_columns"] < hr["populated_columns"]:
                print(f"   yield {hr['sheet']}: {hr['mapped_columns']}/{hr['populated_columns']} "
                      f"populated columns mapped")
    else:
        print(payload)


if __name__ == "__main__":
    main()
