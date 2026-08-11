#!/usr/bin/env python3
"""deliver.py - Stage 7. Assemble the three deliverables.

  1. the dashboard .html (copied into deliverables/ under the project filename)
  2. <slug>_Source_Ledger.xlsx (via ledger.py) - field-level traceability, one row
     per (property, field) -> the source file + locator it came from.
  3. <slug>_Gaps_Report.md - every 'tbd', unmatched asset, conflict and
     enrichment gap, each with a 'how to close it' note. Honest by construction.
  4. <slug>_Longlist.xlsx - a FLAT data view: one property per ROW, variables in
     COLUMNS (the broker-facing table). Sits alongside the Source Ledger, which keeps
     the field-level provenance. Falls back to CSV if openpyxl is unavailable.

CLI:
  python deliver.py --canonical canonical.json --html built.html --ledger ledger.csv \
                    --out-dir deliverables [--slug Normal] [--filename name.html]
"""
from __future__ import annotations

import argparse
import json
import os
import shutil
import sys

import _common as C
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))

MARKER_NAME = ".delivery_complete.json"  # written LAST by main(); see B01 at the write site


def delivery_complete(out_dir) -> bool:
    """True only when a delivery finished AND everything it vouched for is still present.

    The Stage-7 resume guard keys on THIS, not on the dashboard. The dashboard is written
    first, so keying on it meant an incomplete delivery satisfied the guard forever. (B01)"""
    out = Path(out_dir)
    try:
        rec = json.loads((out / MARKER_NAME).read_text(encoding="utf-8"))
    except Exception:
        return False
    names = rec.get("artefacts")
    if not isinstance(names, list) or not names:
        return False
    return all((out / str(n)).exists() for n in names)


CORE = ["warehouseArea", "warehouseRent", "status", "city", "developer", "lat", "lng",
        "clearHeight", "earlyAccess", "motorway"]
CLOSE = {
    "warehouseRent": "request a headline rent from the landlord/agent",
    "warehouseArea": "confirm GLA with the developer",
    "clearHeight": "request the technical spec sheet",
    "lat": "geocode the site or ask for an exact pin",
    "lng": "geocode the site or ask for an exact pin",
    "motorway": "confirm the nearest motorway/corridor",
    "earlyAccess": "confirm the delivery / early-access date",
}


# Fields that are NEVER a broker chase, and why. Three groups, each load-bearing:
#  * INTERNAL / DERIVED - ids, media, the numeric twins of a display string, unit labels.
#  * THE PIPELINE'S OWN "not applicable" MARKERS - `landPrice` and `reit`. merge.py carves
#    both out of unknown-handling (twice, with the identical `field not in ("landPrice",
#    "reit")` line) and _common.fill_render_sentinels writes landPrice = "—". Listing them
#    would report a missing land price and a missing REIT flag for EVERY property in EVERY
#    lease-only longlist - measured: 2 of 11 phantom chases on a well-sourced UK property.
#  * `country`/`region` - ENRICHMENT DERIVES these (enrich.py fills an unknown country from
#    coordinates, with its own ledger trace), so telling a broker to chase an agent for a
#    field the next pass fills is a wrong action.
NOT_CHASEABLE = {
    "id", "photo", "plan", "gallery", "lat", "lng",          # lat/lng stay in CORE instead
    "warehouseAreaVal", "officeAreaVal", "warehouseRentVal", "officeRentVal",
    "areaUnit", "rentUnit", "regionCode", "preBaked", "distances",
    "landPrice", "reit",                                      # "not applicable" markers
    "country", "region",                                      # derived by enrichment
}


def _is_tbd(v):
    """Unknown for Gaps-Report purposes.

    Delegates to the shared sentinel set and ADDS `"??"` locally. `"??"` is what
    _common.REQUIRED_TEXT_SENTINELS writes for an unresolved country and what the chrome's
    own `isAbsent` treats as absent - but normalize.looks_unknown carries `"?"`, NOT `"??"`.
    The widening is deliberately LOCAL: looks_unknown has ~24 call sites and feeds
    _common.core_fill -> record_is_poor -> run.py's vision-routing probe, so widening it
    there would change WHICH INPUT FILES THE LLM IS ASKED TO READ as a side effect of a
    Gaps-Report wording change. Documented divergence, not an oversight."""
    return C._N.looks_unknown(v) or str(v).strip() == "??"


def _chaseable_fields(props: list[dict]) -> list[str]:
    """Every field a broker could actually chase, in a stable order.

    Derived from the fields the DATA carries (union across properties) intersected with the
    schema, minus NOT_CHASEABLE - never from the schema alone, so a field this dataset's
    market does not use is not invented into an action list."""
    try:
        schema = json.loads(Path(C.SCHEMA_FILE).read_text(encoding="utf-8"))
        allowed = set(((schema.get("$defs", {}).get("property", {})
                        or {}).get("properties", {}) or {}).keys())
    except Exception:
        allowed = set()
    seen: set = set()
    for p in props:
        seen |= set(p.keys())
    pool = (seen & allowed) if allowed else seen
    # `*Val` is the derived NUMERIC TWIN of a display string (warehouseAreaVal,
    # expansionParkVal, ...). It is never a broker chase - the string it mirrors already is.
    # Excluded by SUFFIX, not by enumeration, so a future twin cannot leak in.
    return [f for f in sorted(pool - NOT_CHASEABLE - set(CORE)) if not f.endswith("Val")]


def _close_note(field: str) -> str:
    """How to close a gap.

    Bespoke, party-naming advice lives ONLY in CLOSE. Everything else gets a
    PROVENANCE-shaped note, never a party-shaped one: which counterparty holds a given
    figure is a fact about the deal that Python is not entitled to assert (a standing
    second-hand building's counterparty is a landlord, not a developer)."""
    return CLOSE.get(field, "not stated in any source supplied for this property - ask the "
                            "sender if it is decision-relevant")


def gaps_report(canonical: dict, slug: str, work_dir: Path | None = None) -> str:
    props = canonical["properties"]
    meta = canonical.get("meta", {})
    lines = [f"# {slug} - Longlist Gaps Report", "",
             f"Generated {meta.get('generatedAt','')} from {len(props)} properties. "
             "Every item below is a genuine unknown surfaced honestly, not a defect. "
             "Close them with the landlord/agent before the dashboard goes to the client.", ""]

    # per-property tbd core fields
    lines.append("## Missing data by property")
    any_gap = False
    for p in props:
        tbd = [f for f in CORE if _is_tbd(p.get(f))]
        if tbd:
            any_gap = True
            notes = "; ".join(f"`{f}` ({CLOSE.get(f,'confirm with source')})" for f in tbd)
            lines.append(f"- **{p.get('park','?')}** ({p.get('city','?')}, id {p.get('id')}): {notes}")
    if not any_gap:
        lines.append("- None - every property carries all core fields.")
    lines.append("")

    # EVERY OTHER unknown, not just the core ten. Both critical reviewers concluded that
    # listing only CORE is WHY an extraction miss ships invisibly: the card shows `tbd`, the
    # ledger asserts "absent in all sources", and nothing in the delivered pack points at
    # it. Split in two so the action list stays readable:
    #   (1) per-property chases for fields SOME property carries - a real gap, because the
    #       market clearly quotes that attribute;
    #   (2) an INVENTORY of fields NO source carried for any option. These are NOT chases:
    #       the dashboard itself drops such a field (the chrome's FIELD_PRESENT hides a row
    #       no input ever filled), so calling them actions would invert the product's own
    #       rule and bury the real gaps under phantoms.
    chase = _chaseable_fields(props)
    carried = {f for f in chase if any(not _is_tbd(p.get(f)) for p in props)}
    never = [f for f in chase if f not in carried]

    reqs = meta.get("requirements") or {}
    lines.append("## Other missing fields by property")
    if reqs:
        lines.append(f"Ordered by this client's stated requirements first "
                     f"({', '.join(sorted(str(k) for k in reqs)[:8])}).")
    else:
        lines.append("No client requirements were supplied, so this list is the full set of "
                     "attributes the market quoted for at least one option - it is not "
                     "scoped to a brief.")
    _req_first = sorted(carried, key=lambda f: (f not in reqs, f))
    any_other = False
    for p in props:
        tbd = [f for f in _req_first if _is_tbd(p.get(f))]
        if tbd:
            any_other = True
            notes = "; ".join(f"`{f}` ({_close_note(f)})" for f in tbd)
            lines.append(f"- **{p.get('park','?')}** ({p.get('city','?')}, "
                         f"id {p.get('id')}): {notes}")
    if not any_other:
        lines.append("- None - every property carries every attribute the market quoted.")
    lines.append("")

    if never:
        lines.append("## Fields no source provided for any longlist entry")
        lines.append("Not action items: no input carried these for any option, so the "
                     "dashboard hides them entirely. Listed only so you can see what this "
                     "market does not quote - chase one only if the brief calls for it.")
        lines.append("- " + ", ".join(f"`{f}`" for f in never))
        lines.append("")

    # enrichment gaps
    eg = meta.get("enrichmentGaps", [])
    lines.append("## Enrichment gaps")
    lines += ([f"- {g}" for g in eg] if eg else ["- None."])
    lines.append("")

    # unmatched assets. NOTE the `in meta` test rather than a truthiness test: NO code path
    # writes meta.unmatchedAssets today (readers exist here and in gate_runner, a writer never
    # did), so an empty list was indistinguishable from "never computed" and every Gaps Report
    # printed the FALSE affirmative "None - every image bound to a property". Claiming a check
    # that never ran is exactly the kind of unearned assurance this report exists to avoid.
    # When a writer lands (with the inventory->properties reconciliation floor), the affirmative
    # becomes true and this needs no further change.
    lines.append("## Unmatched assets")
    if meta.get("unmatchedAssets"):
        lines += [f"- {a}" for a in meta["unmatchedAssets"]]
    elif "unmatchedAssets" in meta:
        lines.append("- None - every image bound to a property.")
    else:
        lines.append("- Not checked: image-to-property binding is not yet reconciled against the "
                     "input folder, so this run cannot confirm every supplied image was used. "
                     "Compare the folder's images against the cards if that matters.")
    lines.append("")

    # conflicts
    cf = meta.get("conflicts", [])
    lines.append("## Source conflicts")
    lines += ([f"- {c}" for c in cf] if cf else ["- None recorded."])
    lines.append("")

    # I10: notation variants. These are NOT conflicts - two sources stating the same fact in
    # different notation. They are listed so nothing is hidden (the Data Honesty Standard), and
    # kept OUT of "Source conflicts" so that list stays actionable: padding it with "12.5 m vs
    # 12.5" trains a broker to skim past the entries that need a call to the agent.
    nv = meta.get("notationVariants", [])
    lines.append("## Notation variants (same value, stated differently - no action needed)")
    if nv:
        lines.append("Checked and found to denote the same fact - a unit prefix, a date format, a "
                     "qualifier, or a scheme name inside its own full address.")
        lines += [f"- {v}" for v in nv]
    else:
        lines.append("- None recorded.")
    lines.append("")

    # off-spec keys quarantined at the render boundary (v22 Phase 1)
    osp = meta.get("offspec", [])
    lines.append("## Off-spec keys (quarantined provenance/meta - not shown on cards)")
    if osp:
        for e in osp:
            lines.append(f"- property {e.get('property_id')}: `{e.get('key')}` = {e.get('value')} "
                         f"(add to canonical.schema.json + template to display as a real value)")
    else:
        lines.append("- None.")
    lines.append("")

    # B7: brand-new SCALAR fields. Unlike the off-spec keys above these are KEPT and shown on the
    # card - the pipeline deliberately auto-shows any real scalar attribute. They are listed so the
    # drift is visible: `postcode` once shipped on half a longlist's properties while the section
    # above read "None.", because that section only covers quarantined structures.
    nf = meta.get("newFields", [])
    lines.append("## Fields shown but not declared in the schema")
    if nf:
        lines.append("These are real, sourced values and they DO appear on the card - the dashboard "
                     "auto-shows any scalar attribute. They are listed because the schema does not "
                     "declare them, so nothing else would record that. Declare one in "
                     "`templates/canonical.schema.json` to make it first-class.")
        for e in nf:
            lines.append(f"- property {e.get('property_id')}: `{e.get('key')}`"
                         + (f" (from {e.get('source_file')})" if e.get("source_file") else ""))
    else:
        lines.append("- None - every shown field is declared.")
    lines.append("")

    # possible site plans not captured: a page LOOKED plan-ish (classified 'plan' out of band, or
    # carried a plan title) but a precision guard rejected it and no plan bound - surfaced so a
    # genuinely missed plan is visible (never binds a wrong image; honest miss over false plan).
    pnm = meta.get("planNearMiss", [])
    lines.append("## Possible site plans not captured")
    if pnm:
        for e in pnm:
            for pg in e.get("pages", []):
                loc = f"{pg.get('file', '?')} page {int(pg.get('page', 0)) + 1}"
                lines.append(f"- **{e.get('property', '?')}** ({e.get('city', '')}): {loc} - "
                             f"{pg.get('why', '')} (check the deck; if it is the site plan, "
                             f"set __meta.plan_page)")
    else:
        lines.append("- None.")
    lines.append("")

    # ASSUMED UNITS: the source stated a numeric area but never named its unit, so the
    # dataset's dominant unit was applied WITHOUT conversion. Surfaced because the failure is
    # silent and large - a metric figure labelled sq ft is out by 10.76x, and the magnitude
    # cross-check is blind across the whole realistic warehouse range. Chase the source.
    ua = meta.get("unitAssumptions", [])
    if ua:
        lines.append("## Area units assumed (source did not state one)")
        for e in ua:
            lines.append(f"- **{e.get('property', '?')}**: {e.get('field', 'areaUnit')} assumed "
                         f"**{e.get('assumed', '?')}** - {e.get('why', '')}. Confirm the source's "
                         f"own unit; if it differs, the figure needs converting, not relabelling.")
        lines.append("")

    # MANUAL CORRECTIONS (P1-4). Disclosed in the DELIVERABLE, not just in work/: a value a human
    # corrected by hand is not source data, and the Gaps Report is the one document whose job is
    # honesty. A correction that has stopped matching belongs here too - that is the failure this
    # whole mechanism exists to make impossible to miss (it used to be completely silent).
    # B47: options the broker's SOURCE-AUTHORITY answer removed from the longlist. This sits
    # high in the report and names every one: excluding a property from a client's own list is
    # the most consequential thing the pipeline can do to the dataset, and the broker must be
    # able to see exactly what went and why without opening canonical.json.
    excluded = meta.get("excluded") or []
    if excluded:
        lines.append("## Options excluded (not evidenced by your guiding source)")
        lines.append("You told the run which source decides what belongs on this longlist, so "
                     "these options were left OFF. They were found and read normally - nothing "
                     "failed. If any of them should be on the list, re-run and answer the "
                     "source-authority question with 'the union of both'.")
        for e in excluded:
            srcs = ", ".join(e.get("source_files") or []) or "?"
            lines.append(f"- **{e.get('name', '(unnamed option)')}** - {e.get('why', '')} "
                         f"(found in: {srcs})")
        lines.append("")

    ov = meta.get("overrides", {}) or {}
    applied = ov.get("applied") or []
    if applied:
        lines.append("## Manual corrections applied")
        lines.append("Each was applied to the extracted data AFTER extraction and is re-applied on "
                     "every run, so it survives a re-extraction. Every one also has an `override` "
                     "row in the Source Ledger.")
        for e in applied:
            w = e.get("where", {}) or {}
            for f, new in (e.get("set") or {}).items():
                lines.append(f"- **{f}**: `{e.get('old', {}).get(f)}` -> `{new}` "
                             f"({w.get('source_file', '?')} {e.get('locator', '')}) - "
                             f"{e.get('why', '')}"
                             + (f" [verified by {e['verified_by']}]" if e.get("verified_by") else ""))
        lines.append("")
    _rot = [(k, e) for k in ("stale", "ambiguous", "superseded", "invalid")
            for e in (ov.get(k) or [])]
    if _rot:
        lines.append("## Manual corrections that matched NOTHING (stale - fix or delete)")
        lines.append("These were NOT applied, so the data still holds whatever the source said. "
                     "Either correct the `where` block in `work/overrides.json` or delete the "
                     "entry - a correction nobody notices has stopped working is worse than none.")
        for k, e in _rot:
            if isinstance(e, str):          # `invalid` entries are plain reason strings
                lines.append(f"- **{k}**: {e}")
                continue
            lines.append(f"- **{e.get('id', '?')}** ({k}): {e.get('reason', '')} "
                         f"Intended: {e.get('set', {})} - {e.get('why', '')}")
        lines.append("")

    # KNOWN LIMITATIONS: advisory QA findings that were reviewed, judged non-blocking by the
    # reviewer, and NOT fixed within the bounded QA window (one review round + one improvement
    # round - see gate_runner qa-round). This section is what makes the bound honest: before it,
    # the ONLY way to make a finding disappear was to fix it, so a fresh memoryless reviewer with
    # one more debatable layout nit could reopen the gate forever. "Not fixed" now has a
    # delivered home instead of an infinite loop, which is the skill's own stated doctrine -
    # escalate to the Gaps Report rather than loosen a criterion.
    qa_carried: list = []
    if work_dir:
        try:
            import gate_runner as _GR
            qa_carried = _GR.qa_carried(Path(work_dir))
        except Exception:
            qa_carried = []
    if qa_carried:
        lines.append("## Known limitations (reviewed and accepted at QA)")
        for entry in qa_carried:
            lines.append(f"- {entry}")
        lines.append("")

    # photo matches to confirm (run.py writes <work>/photo_doubts.json) - an uncertain
    # brochure<->property pairing shows a placeholder and is surfaced as a yes/no the
    # broker can confirm to pull the photo in
    pd = (Path(work_dir) / "photo_doubts.json") if work_dir else None
    if pd and pd.exists():
        try:
            doubts = json.loads(pd.read_text(encoding="utf-8"))
        except Exception:
            doubts = []
        if doubts:
            lines.append("## Photo matches to confirm")
            lines += [f"- **{d.get('park')}** -> Is this `{d.get('brochure')}`? "
                      f"If yes, confirm and the photo is pulled in immediately"
                      + (f" ({d.get('note')})" if d.get('note') else "") for d in doubts]
            lines.append("")

    # unreadable / skipped input files (run.py writes <work>/unreadable.json) - the
    # honesty standard: a corrupt/encrypted/empty input is a named gap, never a silent drop
    ur = (Path(work_dir) / "unreadable.json") if work_dir else None
    if ur and ur.exists():
        try:
            items = json.loads(ur.read_text(encoding="utf-8"))
        except Exception:
            items = []
        if items:
            lines.append("## Unreadable / skipped input files")
            lines += [f"- **{it.get('file')}**: {it.get('reason')} "
                      f"(re-save or unlock it and re-run to include it)" for it in items]
            lines.append("")

    # extraction yield (run.py writes <work>/yield_report.md when a field-rich
    # spreadsheet yielded a thin parse - surfaced here so it cannot pass silently)
    yr = (Path(work_dir) / "yield_report.md") if work_dir else None
    if yr and yr.exists():
        lines.append("## Extraction yield (unmapped tracker columns)")
        body = [ln for ln in yr.read_text(encoding="utf-8").splitlines()
                if ln.startswith("- ")]
        lines += body or ["- (see yield_report.md in the work folder)"]
        lines.append("")
    return "\n".join(lines)


# The flat Longlist export - one property per ROW, variables in COLUMNS. Field name
# -> friendly header, in a sensible reading order. The two "__" keys are DERIVED:
# the annual rent display string and the monthly equivalent (annual / 12, same
# currency + per-area convention - no FX, no area maths).
LONGLIST_COLUMNS = [
    ("id", "ID"), ("park", "Property / Park"), ("developer", "Developer"),
    ("landlord", "Landlord"),
    ("city", "City"), ("region", "Region"), ("country", "Country"),
    ("status", "Status"), ("permitting", "Permitting"), ("earlyAccess", "Early access"),
    ("warehouseArea", "Warehouse area"), ("areaUnit", "Area unit"),
    ("plotArea", "Plot area"), ("divisibleFrom", "Divisible from"),
    ("officeArea", "Office area"), ("clearHeight", "Clear height"),
    ("floorLoad", "Floor load"), ("sprinklers", "Sprinklers"),
    ("loadingDocks", "Loading docks"), ("overheadDoors", "Overhead doors"),
    ("electricity", "Electricity"), ("truckParking", "Truck parking"),
    ("carParking", "Car parking"),
    ("__rent_annual", "Warehouse rent (annual)"),
    ("__rent_monthly", "Warehouse rent (monthly)"),
    ("__total_annual", "Total annual rent"),
    ("__total_monthly", "Total monthly rent"),
    ("rentUnit", "Rent unit"), ("officeRent", "Office rent"),
    ("serviceCharge", "Service charge"), ("landPrice", "Land price"),
    ("leaseTerm", "Lease term"), ("rentFree", "Rent-free period"),
    # BREEAM and EPC are DIFFERENT certificates and each needs its own column. A single
    # "Certification" column fed only by `breeam` meant a property whose BREEAM is genuinely
    # unstated read "tbd" in the client Excel while its EPC rating - stated twice at source and
    # shipping in canonical.json and the dashboard - appeared nowhere in the workbook at all.
    ("incentives", "Incentives"), ("breeam", "BREEAM"), ("epc", "EPC"),
    ("motorway", "Motorway / corridor"), ("lat", "Latitude"), ("lng", "Longitude"),
    ("mapLink", "Map link"),
]
_WIDE = {"park", "__rent_annual", "__rent_monthly", "__total_annual", "__total_monthly",
         "incentives", "mapLink", "developer", "landlord"}


def _cell(v):
    """Keep numbers numeric (so the sheet sorts), pass strings through, and turn an
    empty/None into an explicit 'tbd' (the honesty standard - never a blank guess)."""
    if v is None:
        return "tbd"
    if isinstance(v, bool):
        return str(v)
    if isinstance(v, str) and v.strip() == "":
        return "tbd"
    return v


def _rent_monthly(p: dict, default_ru: str = "€/sq m/yr") -> str:
    """Monthly headline rent = annual warehouseRentVal / 12, KEPT in its own currency
    and per-area convention. 'tbd' when there is no numeric annual rate to divide."""
    v = p.get("warehouseRentVal")
    if not isinstance(v, (int, float)) or isinstance(v, bool) or v <= 0:
        return "tbd"
    ru = (p.get("rentUnit") or default_ru).split("/")
    cur = (ru[0].strip() if ru and ru[0].strip() else "€")
    per = (ru[1].strip() if len(ru) > 1 and ru[1].strip() else "sq m")
    return f"{cur} {v / 12:.2f} / {per} / mo"


def _total_rent(p: dict, monthly: bool = False) -> str:
    """Total rent = GLA x rate, mirroring the dashboard's totalAnnualRent: split into
    warehouse + office when a separate office rate exists, else the single warehouse
    rate over total GLA (warehouse + office area). 'tbd' when no positive warehouse
    rate/area. Same currency only (no FX); areas are already aligned by merge."""
    wr, wa = p.get("warehouseRentVal"), p.get("warehouseArea")
    if not isinstance(wr, (int, float)) or isinstance(wr, bool) or wr <= 0:
        return "tbd"
    if not isinstance(wa, (int, float)) or isinstance(wa, bool) or wa <= 0:
        return "tbd"
    oa = p.get("officeAreaVal")
    oa = oa if isinstance(oa, (int, float)) and not isinstance(oa, bool) and oa > 0 else 0
    orr = p.get("officeRentVal")
    orr = orr if isinstance(orr, (int, float)) and not isinstance(orr, bool) and orr > 0 else None
    annual = (wa * wr + oa * orr) if (orr is not None and oa > 0) else ((wa + oa) * wr)
    v = annual / 12 if monthly else annual
    cur = ((p.get("rentUnit") or "€/x/yr").split("/")[0] or "€").strip()
    return f"{cur} {round(v):,} / {'mo' if monthly else 'yr'}"


def longlist_xlsx(canonical: dict, out_path: Path) -> None:
    """Write the flat one-property-per-row workbook (CSV fallback if no openpyxl)."""
    props = canonical.get("properties", [])
    meta = canonical.get("meta", {}) or {}
    units = meta.get("units", {}) or {}
    default_ru = units.get("rent") or "€/sq m/yr"
    headers = [h for _, h in LONGLIST_COLUMNS]
    # ASSUMED UNITS ARE DISCLOSED IN THE WORKBOOK, not only in the Gaps Report. (B38)
    #
    # This is the artefact the broker actually forwards, and an "Area unit" of a bare "sq ft"
    # reads as sourced when the source stated nothing - a 10.76x risk presented as fact. The
    # dashboard CARD cannot be fixed from here: its unit is DATASET-wide
    # (`PROPS.find(p => p.areaUnit)`), so no per-property data change touches it; that half is
    # a template item. But this column is already a free string, so the honest label costs no
    # schema change, no record mutation and no chrome edit.
    # Joined by ID, never by park name - two phases of one scheme share a name.
    _assumed_ids = {a.get("id") for a in (meta.get("unitAssumptions") or [])
                    if isinstance(a, dict) and a.get("id") is not None}

    def value_for(p, key):
        if key == "__rent_annual":
            return _cell(p.get("warehouseRent"))
        if key == "__rent_monthly":
            return _rent_monthly(p, default_ru)
        if key == "__total_annual":
            return _total_rent(p, False)
        if key == "__total_monthly":
            return _total_rent(p, True)
        if key == "areaUnit" and p.get("id") in _assumed_ids and p.get("areaUnit"):
            return f"{p['areaUnit']} (assumed - source stated none)"
        return _cell(p.get(key))

    rows = [[value_for(p, key) for key, _ in LONGLIST_COLUMNS] for p in props]
    out = Path(out_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
        wb = Workbook()
        ws = wb.active
        ws.title = "Longlist"
        ws.append(headers)
        hdr_fill = PatternFill("solid", fgColor="003F2D")
        for c in ws[1]:
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = hdr_fill
            c.alignment = Alignment(vertical="center", wrap_text=True)
        for r in rows:
            ws.append(r)
        ws.freeze_panes = "B2"  # freeze the header row + the ID column
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows) + 1}"
        for i, (key, _h) in enumerate(LONGLIST_COLUMNS, start=1):
            ws.column_dimensions[get_column_letter(i)].width = (
                30 if key in _WIDE else 8 if key == "id" else 16)
        tmp = out.with_suffix(out.suffix + ".tmp")
        wb.save(tmp)
        os.replace(tmp, out)
        print(f"OK Longlist -> {out} ({len(rows)} properties x {len(headers)} fields)")
    except Exception as e:
        import csv as _csv
        fallback = out.with_suffix(".csv")
        tmp = fallback.with_suffix(fallback.suffix + ".tmp")
        with open(tmp, "w", newline="", encoding="utf-8") as fh:
            w = _csv.writer(fh)
            w.writerow(headers)
            w.writerows(rows)
        os.replace(tmp, fallback)
        print(f"NOTE openpyxl unavailable ({e}); wrote CSV fallback -> {fallback}")


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--canonical", required=True)
    ap.add_argument("--html", required=True)
    ap.add_argument("--ledger")
    ap.add_argument("--out-dir", required=True)
    ap.add_argument("--slug", default="Longlist")
    ap.add_argument("--filename")
    args = ap.parse_args()

    out = Path(args.out_dir)
    out.mkdir(parents=True, exist_ok=True)
    canonical = json.loads(Path(args.canonical).read_text(encoding="utf-8"))

    # 1. html
    fname = args.filename or f"CBRE_Property_Dashboard_{args.slug}.html"
    dst = out / fname
    tmp = dst.with_suffix(dst.suffix + ".tmp")
    shutil.copyfile(args.html, tmp)
    os.replace(tmp, dst)
    print(f"OK dashboard -> {dst}")

    # 2. ledger - exported IN-PROCESS (same interpreter, no subprocess) so it cannot
    # silently fail on a sandbox where sys.executable can't be re-invoked; the spine
    # captures this stdout in quiet mode. cmd_export already degrades to a .csv copy
    # if openpyxl is missing, so the deliverable is always written.
    if args.ledger and Path(args.ledger).exists():
        import ledger
        try:
            ledger.cmd_export(argparse.Namespace(
                ledger=args.ledger, out=str(out / f"{args.slug}_Source_Ledger.xlsx")))
        except Exception as e:
            print(f"WARNING: Source Ledger export failed: {e}", file=sys.stderr)

    # 3. gaps report (the work dir = the canonical's folder; yield_report.md lives there)
    gaps = out / f"{args.slug}_Gaps_Report.md"
    C.atomic_write_text(gaps, gaps_report(canonical, args.slug,
                                          work_dir=Path(args.canonical).resolve().parent))
    print(f"OK gaps report -> {gaps}")

    # 4. flat longlist workbook (one property per row, variables in columns) - a
    # broker-facing data view alongside the field-level Source Ledger. Guarded so a
    # workbook hiccup can never block the dashboard hand-off.
    try:
        longlist_xlsx(canonical, out / f"{args.slug}_Longlist.xlsx")
    except Exception as e:
        print(f"WARNING: Longlist export failed: {e}", file=sys.stderr)

    # 5. THE COMPLETION MARKER - written LAST, and the only thing that means "delivered".
    #
    # The four artefacts above are each atomic; the SET was not. Stage 7's resume guard used
    # to key on the DASHBOARD, which lands FIRST, so a kill in steps 2-4 left the guard
    # satisfied by an incomplete delivery: either three artefacts never appeared and every
    # later run resume-skipped past them (final_gate then failing with no remediation, and
    # Stage 7 exits 0 so nothing bounded the loop), or - worse, because it ships - a v2
    # dashboard sat beside a v1 Gaps Report and Longlist and every presence check passed.
    #
    # The marker names what it vouches for, so a later deletion is detectable. It does NOT
    # require the Longlist to exist: that export is deliberately allowed to fail (above), and
    # a predicate that demanded it would be unsatisfiable on a box where openpyxl is broken -
    # deliver would then re-run on every single pass forever while final_gate still blocked.
    # That would trade one unbounded loop for another. (B01)
    _vouched = [fname, f"{args.slug}_Gaps_Report.md"]
    if args.ledger and Path(args.ledger).exists():
        _led = out / f"{args.slug}_Source_Ledger.xlsx"
        _vouched.append(_led.name if _led.exists() else f"{args.slug}_Source_Ledger.csv")
    C.atomic_write_text(out / MARKER_NAME, json.dumps(
        {"schema_version": 1, "slug": args.slug, "artefacts": _vouched},
        ensure_ascii=False, indent=2))
    print(f"OK delivery complete -> {out / MARKER_NAME}")


if __name__ == "__main__":
    main()
