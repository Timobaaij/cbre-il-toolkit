#!/usr/bin/env python3
"""ledger.py - the Source Ledger (field-level traceability).

Mirrors the cbre-il-account-briefing ledger: every populated property field (and
every explicit 'tbd') is one row tying the value to the input file + locator it
came from. Subcommands:
  validate <ledger.csv>            : reject rows missing a required column
  export   <ledger.csv> --out x.xlsx : write a formatted .xlsx (frozen header,
                                       autofilter, column widths). Falls back to a
                                       copy of the CSV if openpyxl is unavailable.

Columns (order fixed):
  property_id, record_type, field, value, source_file, source_locator,
  source_type, extractor, confidence, conflict_note, verified
Required at merge: field, value, source_file, source_locator, source_type.
"""
from __future__ import annotations

import argparse
import csv
import os
import sys
from pathlib import Path

import _common as C

COLUMNS = ["property_id", "record_type", "field", "value", "source_file",
           "source_locator", "source_type", "extractor", "confidence",
           "conflict_note", "verified"]
REQUIRED = ["field", "value", "source_file", "source_locator", "source_type"]


def read_rows(path: Path) -> list[dict]:
    with open(path, newline="", encoding="utf-8") as fh:
        return list(csv.DictReader(fh))


def cmd_validate(args) -> int:
    rows = read_rows(Path(args.ledger))
    if not rows:
        print("[FAIL] ledger has no data rows - every property field must trace to a source (S4-52)")
        print("STATUS: BLOCKED (empty ledger)")
        return 1
    bad = []
    for i, r in enumerate(rows, start=2):  # +1 header, +1 1-based
        missing = [c for c in REQUIRED if not str(r.get(c, "")).strip()]
        if missing:
            bad.append(f"row {i}: missing {missing} (field={r.get('field')!r})")
    if bad:
        for b in bad[:30]:
            print(f"[FAIL] {b}")
        print(f"STATUS: BLOCKED ({len(bad)} incomplete rows)")
        return 1
    print(f"[PASS] ledger complete ({len(rows)} rows, every row traces to a source)")
    print("STATUS: ALL-PASS")
    return 0


def cmd_export(args) -> int:
    rows = read_rows(Path(args.ledger))
    out = Path(args.out)
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = Workbook()
        ws = wb.active
        ws.title = "Source Ledger"
        ws.append(COLUMNS)
        hdr_fill = PatternFill("solid", fgColor="003F2D")
        for c in ws[1]:
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = hdr_fill
            c.alignment = Alignment(vertical="center")
        for r in rows:
            ws.append([r.get(c, "") for c in COLUMNS])
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{chr(64 + len(COLUMNS))}{len(rows) + 1}"
        widths = {"A": 10, "B": 11, "C": 16, "D": 40, "E": 30, "F": 22,
                  "G": 11, "H": 10, "I": 11, "J": 26, "K": 9}
        for col, w in widths.items():
            ws.column_dimensions[col].width = w
        out.parent.mkdir(parents=True, exist_ok=True)
        tmp = out.with_suffix(out.suffix + ".tmp")
        wb.save(tmp)
        os.replace(tmp, out)
        print(f"OK Source Ledger -> {out} ({len(rows)} rows)")
    except Exception as e:
        fallback = out.with_suffix(".csv")
        C.atomic_write_text(fallback, Path(args.ledger).read_text(encoding="utf-8"))
        print(f"NOTE openpyxl unavailable ({e}); wrote CSV fallback -> {fallback}")
    return 0


def main() -> None:
    ap = argparse.ArgumentParser()
    sub = ap.add_subparsers(dest="cmd", required=True)
    p = sub.add_parser("validate"); p.add_argument("ledger"); p.set_defaults(fn=cmd_validate)
    p = sub.add_parser("export"); p.add_argument("ledger"); p.add_argument("--out", required=True)
    p.set_defaults(fn=cmd_export)
    args = ap.parse_args()
    sys.exit(args.fn(args))


if __name__ == "__main__":
    main()
