#!/usr/bin/env python3
"""fixture_test.py - offline end-to-end fixture: the pipeline must agree with its
own gates ON THE FIRST PASS.

Locks in the producer-vs-gate contract (the class of defect behind the "agentic
loops of feedback" complaint): every state merge/match/enrich deterministically
produces must clear every mechanical gate, with no orchestrator intervention.

Covers, with assertions:
  P1  a numeric-only rent (xlsx/vision style) -> derived display inherits prov
      -> trace-coverage ALL-PASS (was: guaranteed "possible fabrication" block)
  P2  a same-file summary-row + detail-page restatement -> merged, so the
      coverage duplicate check cannot block the skill's own output
  P3  two regions sharing a real identical statistic -> enrichment gate NOTES it,
      never blocks (the only "fix" for a block on true data is falsifying it)
  P4  sentinel identity fields (country '??' etc.) get their gap ledger rows
  P7  vision-transcribed fields carry confidence Medium, not a hardcoded High
  P8  coverage does not demand lat/lng when geocoding was not requested
  V1-3 final_gate verdict protocol: amber passes, prose "verdict: red ..." in a
      reviewer's reasoning never false-blocks, a missing VERDICT line FAILS
Plus negatives proving the gates still bite: a populated field with no ledger
row must BLOCK trace-coverage; a reviewers dir with a malformed verdict must
fail the final gate.

Run: python evals/fixture_test.py     (exit 0 on success, 1 on any failure)
Offline by design - no network, no Playwright, no real brochures.
"""
from __future__ import annotations

import csv
import io
import json
import sys
import tempfile
from contextlib import redirect_stdout
from pathlib import Path

HELPERS = Path(__file__).resolve().parent.parent / "helpers"
sys.path.insert(0, str(HELPERS))
import build_dashboard  # noqa: E402
import deliver          # noqa: E402
import final_gate       # noqa: E402
import gate_runner      # noqa: E402
import ledger as L      # noqa: E402
import merge            # noqa: E402
import i18n as I18N     # noqa: E402

FAILS: list[str] = []


def check(ok: bool, label: str) -> None:
    print(f"  [{'PASS' if ok else 'FAIL'}] {label}")
    if not ok:
        FAILS.append(label)


def call(module, *cmd, quiet=True) -> int:
    """Run a helper main() in-process (the run.py pattern); return its exit code."""
    saved = sys.argv
    sys.argv = [getattr(module, "__name__", "helper"), *[str(c) for c in cmd]]
    buf = io.StringIO()
    try:
        if quiet:
            with redirect_stdout(buf):
                module.main()
        else:
            module.main()
        rc = 0
    except SystemExit as e:
        rc = e.code if isinstance(e.code, int) else (0 if e.code is None else 1)
    except Exception as e:
        print(f"    (crash: {type(e).__name__}: {e})")
        rc = 1
    finally:
        sys.argv = saved
    return rc


# --------------------------------------------------------------------------- #
def _px(rgb) -> str:
    """A tiny DISTINCT PHOTO-kind data URI per property: the images gate now blocks a
    longlist that is mostly placeholders, shares one identical hero, OR whose hero is a
    non-photo (map/plan/screenshot), so the healthy fixture carries real, distinct,
    PHOTOGRAPHIC heroes. A random-tone texture (NOT a solid - a solid is a 'logo') reads
    as 'photo' (high luminance entropy, no flat fills); seeded by rgb so the SAME rgb is
    byte-identical (the duplicate-hero test relies on that)."""
    import base64
    import io as _io
    import random as _r
    from PIL import Image
    rnd = _r.Random(str(rgb))
    img = Image.new("RGB", (64, 48))
    img.putdata([(rnd.randint(0, 255), rnd.randint(0, 255), rnd.randint(0, 255))
                 for _ in range(64 * 48)])
    buf = _io.BytesIO()
    img.save(buf, format="PNG")
    return "data:image/png;base64," + base64.b64encode(buf.getvalue()).decode("ascii")


def fixture_records(work: Path) -> list[Path]:
    """Three record files exercising the producer-vs-gate seams."""
    brochure = [
        {  # detail page
            "park": "Alpha Park", "developer": "CTP", "city": "Pilsen", "country": "CZ",
            "status": "Existing", "warehouseArea": 40000,
            "warehouseRent": "€60 / sq m / year", "warehouseRentVal": 60.0,
            "clearHeight": "12 m", "photo": _px((20, 90, 60)),
            "__meta": {"source_file": "Brochure A.pdf", "source_type": "pdf",
                       "locator_base": "page 3",
                       "prov": {"park": "page 3", "developer": "page 3", "city": "page 3",
                                "country": "page 3", "status": "page 3",
                                "warehouseArea": "page 3", "warehouseRent": "page 3",
                                "warehouseRentVal": "page 3", "clearHeight": "page 3"}},
        },
        {  # P2: same-file SUMMARY-ROW restatement of the same unit (same key + area)
            "park": "Alpha Park", "developer": "CTP", "city": "Pilsen", "country": "CZ",
            "status": "Existing", "warehouseArea": 40000,
            "__meta": {"source_file": "Brochure A.pdf", "source_type": "pdf",
                       "locator_base": "page 2 (summary table)",
                       "prov": {"park": "page 2", "warehouseArea": "page 2"}},
        },
        {  # same park, materially different area = a distinct phase, stays its own card
            "park": "Alpha Park", "developer": "CTP", "city": "Pilsen", "country": "CZ",
            "status": "Under construction", "warehouseArea": 80000, "photo": _px((40, 70, 110)),
            "warehouseRent": "€58 / sq m / year", "warehouseRentVal": 58.0,
            "__meta": {"source_file": "Brochure A.pdf", "source_type": "pdf",
                       "locator_base": "page 5",
                       "prov": {"park": "page 5", "developer": "page 5", "city": "page 5",
                                "country": "page 5", "status": "page 5",
                                "warehouseArea": "page 5", "warehouseRent": "page 5",
                                "warehouseRentVal": "page 5"}},
        },
    ]
    tracker = [
        {  # P1: NUMERIC-ONLY rent - canonicalize() derives the display string
            "park": "Beta Park", "developer": "Panattoni", "city": "Brno", "country": "CZ",
            "status": "Speculative", "warehouseArea": 25000, "warehouseRentVal": 55.0,
            "photo": _px((120, 60, 30)),
            "__meta": {"source_file": "Tracker.xlsx", "source_type": "xlsx",
                       "locator_base": "Sheet1!r4",
                       "prov": {"park": "Sheet1!r4", "developer": "Sheet1!r4",
                                "city": "Sheet1!r4", "country": "Sheet1!r4",
                                "status": "Sheet1!r4", "warehouseArea": "Sheet1!r4",
                                "warehouseRentVal": "Sheet1!r4"}},
        },
    ]
    vision = [
        {  # P4 + P7: vision transcription, country unreadable -> sentinel gap row
            "park": "Gamma Park", "developer": "VGP", "city": "Ostrava",
            "status": "Existing", "warehouseArea": 18000, "photo": _px((80, 80, 20)),
            "warehouseRent": "€52 / sq m / year", "warehouseRentVal": 52.0,
            "__meta": {"source_file": "Scan B.pdf", "source_type": "pdf",
                       "locator_base": "page 1",
                       "prov": {"park": "page 1 (vision transcription)",
                                "developer": "page 1 (vision transcription)",
                                "city": "page 1 (vision transcription)",
                                "status": "page 1 (vision transcription)",
                                "warehouseArea": "page 1 (vision transcription)",
                                "warehouseRent": "page 1 (vision transcription)",
                                "warehouseRentVal": "page 1 (vision transcription)"}},
        },
    ]
    out = []
    for name, recs in (("brochure.json", brochure), ("tracker.json", tracker),
                       ("vision.json", vision)):
        f = work / name
        f.write_text(json.dumps(recs, ensure_ascii=False), encoding="utf-8")
        out.append(f)
    return out


def main() -> int:
    work = Path(tempfile.mkdtemp(prefix="cbre_longlist_fixture_"))
    src = work / "inputs"; src.mkdir()
    canonical = work / "canonical.json"
    ledger_csv = work / "source_ledger.csv"
    built = work / "built.html"
    records = fixture_records(work)

    print("Stage 2 - merge (same-file restatement, numeric-only rent, vision prov):")
    rc = call(merge, "--records", *records, "--source-dir", src,
              "--out", canonical, "--ledger", ledger_csv)
    check(rc == 0, "merge completes")
    data = json.loads(canonical.read_text(encoding="utf-8"))
    # 4 = Alpha detail (+ its summary row MERGED in), Alpha phase II (distinct,
    # different area), Beta, Gamma. 5 would mean the restatement shipped as a dup.
    check(len(data["properties"]) == 4,
          f"P2: summary row merged into its detail page (4 properties, got {len(data['properties'])})")
    alphas_40k = [p for p in data["properties"]
                  if p["park"] == "Alpha Park" and p.get("warehouseArea") == 40000]
    check(len(alphas_40k) == 1 and alphas_40k[0].get("clearHeight") == "12 m",
          "P2: merged record carries the detail page's fields, no duplicate card")
    beta = next((p for p in data["properties"] if p["park"] == "Beta Park"), {})
    check(beta.get("warehouseRent", "").startswith("€55"),
          "P1: display rent derived from the numeric")

    rows = list(csv.DictReader(open(ledger_csv, newline="", encoding="utf-8")))
    beta_rent_rows = [r for r in rows if r["field"] == "warehouseRent"
                      and r["source_file"] == "Tracker.xlsx"]
    check(bool(beta_rent_rows) and "derived from warehouseRentVal" in beta_rent_rows[0]["source_locator"],
          "P1: derived display rent has a ledger row inheriting the numeric's source")
    gamma_rows = [r for r in rows if "vision transcription" in r.get("source_locator", "")]
    check(bool(gamma_rows) and all(r["confidence"] == "Medium" for r in gamma_rows),
          "P7: vision-transcribed fields are confidence Medium")
    country_gap = [r for r in rows if r["field"] == "country" and r["source_type"] == "gap"]
    check(bool(country_gap), "P4: sentinel country ('??') has its gap row")

    # Phase 2: the REAL merge.main --ui-overrides bake (not a re-implemented filter). A
    # fallback cache with every EN key translated + the _en_sha meta + a stray non-EN key
    # must yield a canonical whose meta.ui_overrides is EXACTLY the EN key set.
    ui_cache = work / "da_cache.json"
    _fake = {k: ("DA_" + v if isinstance(v, str) else v) for k, v in I18N.EN.items()}
    _fake["kpi_wh_area_sub_fmt"] = "{area} DA"; _fake["kpi_rent_sub_fmt"] = "DA {unit}"
    _fake["_en_sha"] = I18N.en_sha(); _fake["STRAY_DATA_KEY"] = "must not enter canonical"
    ui_cache.write_text(json.dumps(_fake, ensure_ascii=False), encoding="utf-8")
    canonical_ui = work / "canonical_ui.json"
    rc_ui = call(merge, "--records", *records, "--source-dir", src,
                 "--out", canonical_ui, "--ledger", work / "ledger_ui.csv",
                 "--language", "Danish", "--ui-overrides", ui_cache)
    check(rc_ui == 0, "merge --ui-overrides completes")
    meta_ui = json.loads(canonical_ui.read_text(encoding="utf-8")).get("meta", {})
    ov = meta_ui.get("ui_overrides", {})
    check(set(ov) == set(I18N.EN) and "_en_sha" not in ov and "STRAY_DATA_KEY" not in ov,
          "merge --ui-overrides bakes meta.ui_overrides = EXACTLY the EN key set (stray/_en_sha dropped)")
    check(ov.get("tab_grid") == "DA_Grid" and meta_ui.get("language") == "Danish",
          "baked meta.ui_overrides carries the translated values + meta.language")

    # ----------------------------------------------------------------------- #
    # #4 - cross-source VALUE-conflict adjudication (LLM override on the exit-10
    # dispatch; the fixed precedence is the DEFAULT + the offline fallback; the
    # field plausibility gate vetoes an implausible pick).
    print("#4 - cross-source value-conflict adjudication (precedence default; gate-verified LLM override):")
    import match as MATCH

    def _conflict_recs(broc_rent, trk_rent):
        """One property described by a brochure (pdf) and a tracker (xlsx) that AGREE on
        identity/area (so they cluster 'auto') but DISAGREE on warehouseRentVal - a genuine
        cross-source value conflict. For a COMMERCIAL field the precedence default is the
        tracker (xlsx outranks pdf), so the non-default candidate is the brochure rent."""
        broc = {
            "park": "Delta Park", "developer": "Prologis", "city": "Lodz", "country": "PL",
            "status": "Existing", "warehouseArea": 30000, "warehouseRentVal": broc_rent,
            "warehouseRent": f"€{broc_rent:g} / sq m / year", "photo": _px((33, 99, 66)),
            "__meta": {"source_file": "Delta brochure.pdf", "source_type": "pdf",
                       "date": "2025-01-01", "locator_base": "page 2",
                       "prov": {"park": "page 2", "developer": "page 2", "city": "page 2",
                                "country": "page 2", "status": "page 2", "warehouseArea": "page 2",
                                "warehouseRentVal": "page 2", "warehouseRent": "page 2"}},
        }
        trk = {
            "park": "Delta Park", "developer": "Prologis", "city": "Lodz", "country": "PL",
            "status": "Existing", "warehouseArea": 30000, "warehouseRentVal": trk_rent,
            "warehouseRent": f"€{trk_rent:g} / sq m / year",
            "__meta": {"source_file": "Tracker.xlsx", "source_type": "xlsx",
                       "date": "2025-02-01", "locator_base": "Sheet1!r9",
                       "prov": {"park": "Sheet1!r9", "developer": "Sheet1!r9", "city": "Sheet1!r9",
                                "country": "Sheet1!r9", "status": "Sheet1!r9",
                                "warehouseArea": "Sheet1!r9", "warehouseRentVal": "Sheet1!r9",
                                "warehouseRent": "Sheet1!r9"}},
        }
        return [broc, trk]

    # (d) conflict_candidates: deterministic, order-independent, GENUINE conflicts only.
    cd_recs = _conflict_recs(70.0, 55.0)  # brochure 70, tracker 55 -> a genuine rent conflict
    cl = MATCH.dedupe(cd_recs)
    check(len(cl) == 1, f"#4d: the brochure + tracker cluster as ONE property (got {len(cl)})")
    cc = merge.conflict_candidates(cl)
    rent_cc = [c for c in cc if c["field"] == "warehouseRentVal"]
    check(len(rent_cc) == 1 and len(rent_cc[0]["candidates"]) == 2,
          "#4d: a genuine 2-source rent conflict is enumerated (2 candidates)")
    # the precedence default for a COMMERCIAL field is the tracker (xlsx > pdf)
    cdef = rent_cc[0]
    default_val = next(c["value"] for c in cdef["candidates"] if c["label"] == cdef["default"])
    check(default_val == 55.0, f"#4d: the precedence default is the tracker rent 55 (got {default_val})")
    # order-independent id: reversing the record order yields the SAME conflict_id
    cc_rev = merge.conflict_candidates(MATCH.dedupe(list(reversed(cd_recs))))
    rent_rev = [c for c in cc_rev if c["field"] == "warehouseRentVal"][0]
    check(rent_rev["conflict_id"] == cdef["conflict_id"],
          "#4d: conflict_id is order-independent (id(a,b) == id(b,a))")
    # the helper keys on match_key+field+sorted-values directly (content-only)
    check(merge.conflict_id("k|d|p", "warehouseRentVal", [55.0, 70.0])
          == merge.conflict_id("k|d|p", "warehouseRentVal", [70.0, 55.0]),
          "#4d: conflict_id is independent of the value order passed in")
    # an AGREEING field (both say 30000) is NOT a conflict
    check(not any(c["field"] == "warehouseArea" for c in cc),
          "#4d: an agreeing field (warehouseArea 30000) is NOT enumerated")

    cwork = work / "conflict"; cwork.mkdir()
    csrc = cwork / "inputs"; csrc.mkdir()
    crecs = cwork / "delta.json"
    crecs.write_text(json.dumps(cd_recs, ensure_ascii=False), encoding="utf-8")

    def _merge_conflict(out_name, field_dec=None):
        outp = cwork / out_name
        cmd = ["--records", crecs, "--source-dir", csrc, "--out", outp]
        if field_dec is not None:
            fdf = cwork / (out_name + ".fd.json")
            fdf.write_text(json.dumps(field_dec, ensure_ascii=False), encoding="utf-8")
            cmd += ["--field-decisions", fdf]
        rc = call(merge, *cmd)
        return rc, json.loads(outp.read_text(encoding="utf-8"))

    # (a) OFFLINE (no --field-decisions): the precedence winner is chosen AND the conflict
    # is recorded - the existing behaviour, unchanged.
    rc_a, data_a = _merge_conflict("canonical_a.json")
    pa = data_a["properties"][0]
    check(rc_a == 0 and pa.get("warehouseRentVal") == 55.0,
          f"#4a: offline, the precedence default (tracker 55) wins (got {pa.get('warehouseRentVal')})")
    check(any("warehouseRentVal" in c for c in data_a["meta"]["conflicts"]),
          "#4a: the discarded brochure rent is recorded in meta.conflicts")

    # (b) an LLM pick of the NON-default candidate (the brochure rent 70, gate-passing) ->
    # merge OVERRIDES + records the rationale.
    # the non-default label is whichever candidate is NOT the default
    nondef_label = next(c["label"] for c in cdef["candidates"] if c["label"] != cdef["default"])
    rc_b, data_b = _merge_conflict("canonical_b.json",
                                   {cdef["conflict_id"]: {"pick": nondef_label,
                                                          "reason": "newer email confirms €70"}})
    pb = data_b["properties"][0]
    check(rc_b == 0 and pb.get("warehouseRentVal") == 70.0,
          f"#4b: the gate-passing LLM pick OVERRIDES precedence (70, got {pb.get('warehouseRentVal')})")
    check(any("LLM override" in c and "warehouseRentVal" in c for c in data_b["meta"]["conflicts"]),
          "#4b: the override + rationale is recorded in meta.conflicts (auditable, never silent)")
    # the override carries the picked source's prov (selection-only -> still traces) and
    # the display rent is regenerated from the overridden numeric (pair-consistency holds)
    check(pb.get("warehouseRent", "").startswith("€70"),
          "#4b: the display rent is regenerated from the overridden numeric (pair-consistent)")

    # (c) a pick whose VALUE fails the rent plausibility band -> merge IGNORES it, keeps
    # precedence (the gate veto). 5000 €/m²/yr is far above the 500 ceiling.
    bad_recs = _conflict_recs(5000.0, 55.0)  # brochure 5000 is implausible
    bad_cl = MATCH.dedupe(bad_recs)
    bad_cc = [c for c in merge.conflict_candidates(bad_cl) if c["field"] == "warehouseRentVal"][0]
    bad_nondef = next(c["label"] for c in bad_cc["candidates"] if c["label"] != bad_cc["default"])
    bwork = cwork / "delta_bad.json"
    bwork.write_text(json.dumps(bad_recs, ensure_ascii=False), encoding="utf-8")
    bfd = cwork / "bad.fd.json"
    bfd.write_text(json.dumps({bad_cc["conflict_id"]: {"pick": bad_nondef, "reason": "x"}}),
                   encoding="utf-8")
    bout = cwork / "canonical_c.json"
    rc_c = call(merge, "--records", bwork, "--source-dir", csrc, "--out", bout,
                "--field-decisions", bfd)
    data_c = json.loads(bout.read_text(encoding="utf-8"))
    pc = data_c["properties"][0]
    check(rc_c == 0 and pc.get("warehouseRentVal") == 55.0,
          f"#4c: an implausible pick (5000) is VETOED; precedence 55 stands (got {pc.get('warehouseRentVal')})")
    check(any("rejected" in c and "gate" in c for c in data_c["meta"]["conflicts"]),
          "#4c: the gate veto is noted in meta.conflicts")

    # (e) merge.main with NO --field-decisions is byte-identical to today's offline merge
    # (the offline fallback): a second offline run reproduces the same canonical bytes.
    rc_e, data_e = _merge_conflict("canonical_e.json")
    check(json.dumps(data_e, sort_keys=True) == json.dumps(data_a, sort_keys=True),
          "#4e: two offline merges (no --field-decisions) are byte-identical (deterministic fallback)")
    # and a decision that picks the DEFAULT is a no-op (precedence already chose it)
    rc_nd, data_nd = _merge_conflict("canonical_nd.json",
                                     {cdef["conflict_id"]: {"pick": cdef["default"], "reason": "agree"}})
    check(data_nd["properties"][0].get("warehouseRentVal") == 55.0,
          "#4e: a pick of the default candidate is a no-op (precedence stands)")

    # ----------------------------------------------------------------------- #
    # I1 (v22 Phase 1 fix) - end-to-end AUDIT of the off-spec quarantine: the
    # "never silently destroy data" invariant. A stray TOP-LEVEL object (the
    # leak shape - a provenance/meta map that landed on the record instead of
    # under __meta) must never become a displayed field, but it must also never
    # vanish: it has to (a) be pulled out of the property, (b) get its OWN
    # ledger row, and (c) surface in the Gaps Report. A genuine brand-new
    # SCALAR attribute sitting right next to it must survive untouched
    # (auto-show is preserved).
    print("I1 - off-spec quarantine reaches the ledger AND the Gaps Report:")
    owork = work / "offspec_audit"; owork.mkdir()
    osrc = owork / "inputs"; osrc.mkdir()
    offspec_rec = [{
        "park": "Epsilon Park", "developer": "7R", "city": "Wroclaw", "country": "PL",
        "status": "Existing", "warehouseArea": 22000,
        "commune": "Katy Wroclawskie",  # genuine brand-new scalar - must be KEPT
        "prov": {"city": "page 9 (text interpretation)",                    # off-spec
                 "warehouseArea": "page 9 (SUPERFICIES nave total)"},       # LEAK object
        "__meta": {"source_file": "offspec_audit_deck.pdf", "source_type": "pdf",
                   "locator_base": "page 9",
                   "prov": {"park": "page 9", "developer": "page 9", "city": "page 9",
                            "country": "page 9", "status": "page 9", "warehouseArea": "page 9"}},
    }]
    offspec_records = owork / "offspec.json"
    offspec_records.write_text(json.dumps(offspec_rec, ensure_ascii=False), encoding="utf-8")
    offspec_canonical = owork / "canonical.json"
    offspec_ledger = owork / "source_ledger.csv"
    rc_os = call(merge, "--records", offspec_records, "--source-dir", osrc,
                 "--out", offspec_canonical, "--ledger", offspec_ledger)
    check(rc_os == 0, "I1: merge completes on a record carrying an off-spec top-level object")
    data_os = json.loads(offspec_canonical.read_text(encoding="utf-8"))
    props_os = data_os["properties"]
    check(len(props_os) == 1 and "prov" not in props_os[0],
          "I1: the off-spec top-level object never reaches the property (quarantined, not a field)")
    check(props_os[0].get("commune") == "Katy Wroclawskie",
          "I1: a genuine brand-new scalar attribute is KEPT (auto-show preserved)")
    check(any(e.get("key") == "prov" for e in data_os["meta"].get("offspec", [])),
          "I1: meta.offspec records the quarantined key")

    os_rows = list(csv.DictReader(open(offspec_ledger, newline="", encoding="utf-8")))
    check(any(r["record_type"] == "offspec" and r["field"] == "prov" for r in os_rows),
          "I1: the off-spec key has its own audit ledger row (never silently dropped)")

    os_gaps = deliver.gaps_report(data_os, "OffspecAudit")
    check("Off-spec keys" in os_gaps and "prov" in os_gaps,
          "I1: the Gaps Report lists the quarantined off-spec key")

    print("Stage 4 - pre-build mechanical gates (must ALL-PASS first try):")
    for cmd in (("validate-data", canonical), ("self-check",),
                ("coverage", canonical),
                ("trace-coverage", canonical, "--ledger", ledger_csv),
                ("images", canonical), ("enrichment", canonical)):
        check(call(gate_runner, *cmd) == 0, f"gate {cmd[0]} ALL-PASS (P8: no lat/lng demanded sans geocode)"
              if cmd[0] == "coverage" else f"gate {cmd[0]} ALL-PASS")
    check(call(L, "validate", ledger_csv) == 0, "ledger validate ALL-PASS")

    print("P3 - identical real regional statistics must NOTE, not block:")
    twin = json.loads(canonical.read_text(encoding="utf-8"))
    twin["meta"]["enrichment"] = {"regions": True}
    twin["regions"] = {
        "PLZ": {"name": "Pilsen", "country": "CZ", "unemployment": 3.2,
                "unemploymentAsOf": "2025", "sources": "CZSO 2025"},
        "BRN": {"name": "Brno", "country": "CZ", "unemployment": 3.2,
                "unemploymentAsOf": "2025", "sources": "CZSO 2025"},
    }
    twin_path = work / "canonical_regions.json"
    twin_path.write_text(json.dumps(twin, ensure_ascii=False), encoding="utf-8")
    check(call(gate_runner, "enrichment", twin_path) == 0,
          "P3: enrichment gate passes two regions sharing unemployment=3.2 (advisory note)")

    print("Negative - the gates must still bite:")
    fab = json.loads(canonical.read_text(encoding="utf-8"))
    fab["properties"][0]["motorway"] = "D5"  # populated, but NO ledger row
    fab_path = work / "canonical_fabricated.json"
    fab_path.write_text(json.dumps(fab, ensure_ascii=False), encoding="utf-8")
    check(call(gate_runner, "trace-coverage", fab_path, "--ledger", ledger_csv) != 0,
          "trace-coverage BLOCKS a populated field with no ledger row")

    # duplicate-hero check: >=3 properties sharing ONE image is a harvest failure
    import hashlib
    import images as IMG_
    dup_dir = work / "neg_dup"; dup_dir.mkdir()
    shared = _px((10, 120, 80))
    dup_path = dup_dir / "canonical.json"
    dup_path.write_text(json.dumps({"meta": {}, "properties": [
        {"id": n, "park": f"P{n}", "developer": "D", "city": "X", "country": "CZ",
         "status": "Existing", "photo": shared} for n in (1, 2, 3)]}), encoding="utf-8")
    check(call(gate_runner, "images", dup_path) != 0,
          "images gate BLOCKS 3 properties sharing one identical hero")
    h12 = hashlib.sha1(shared.encode("ascii")).hexdigest()[:12]
    (dup_dir / "placeholder_audit_ack.json").write_text(
        json.dumps({"duplicate_photos_ok": [h12]}), encoding="utf-8")
    check(call(gate_runner, "images", dup_path) == 0,
          "the recorded G-images sign-off unblocks a genuine duplicate")

    # NON-PHOTO HERO check (the linchpin): a map / plan / slide-screenshot hero must BLOCK -
    # the independent G-images reviewer flagged exactly this on a real run, but the gate
    # only ADVISED, so it shipped. Now it blocks until the hero is a photo OR a reviewer
    # records nonphoto_hero_ok; a real PHOTO hero passes clean.
    import io as _io2, base64 as _b64
    from PIL import Image as _Img
    np_dir = work / "neg_nonphoto"; np_dir.mkdir()
    _solid = _Img.new("RGB", (64, 48), (210, 205, 180))  # flat fill -> non-photo ('logo')
    _b = _io2.BytesIO(); _solid.save(_b, "PNG")
    nonphoto = "data:image/png;base64," + _b64.b64encode(_b.getvalue()).decode("ascii")
    check(IMG_.classify_data_uri(nonphoto) != "photo",
          "classify_data_uri tags a flat-fill hero as non-photo")
    check(IMG_.classify_data_uri(_px((30, 110, 70))) == "photo",
          "classify_data_uri tags a photographic (texture) hero as photo")
    np_path = np_dir / "canonical.json"
    np_path.write_text(json.dumps({"meta": {}, "properties": [
        {"id": 1, "park": "P1", "developer": "D", "city": "X", "country": "CZ",
         "status": "Existing", "photo": _px((1, 2, 3))},
        {"id": 2, "park": "P2", "developer": "D", "city": "Y", "country": "CZ",
         "status": "Existing", "photo": nonphoto}]}), encoding="utf-8")
    check(call(gate_runner, "images", np_path) != 0,
          "images gate BLOCKS a non-photo (map/plan/screenshot) hero")
    (np_dir / "placeholder_audit_ack.json").write_text(
        json.dumps({"nonphoto_hero_ok": ["2"]}), encoding="utf-8")
    check(call(gate_runner, "images", np_path) == 0,
          "the recorded nonphoto_hero_ok sign-off unblocks the reviewed non-photo hero")

    # P1-6: the placeholder-rate check is IMAGE-SOURCE-AWARE.
    # (a) a record/tracker-only run (NO brochure examined -> meta.placeholderAudit empty)
    #     SHIPS with honest placeholders - never block a bare-spreadsheet dashboard.
    ph = IMG_.placeholder()
    rate_dir = work / "neg_rate"; rate_dir.mkdir()
    rate_path = rate_dir / "canonical.json"
    rate_path.write_text(json.dumps({"meta": {}, "properties": [
        {"id": n, "park": f"P{n}", "developer": "D", "city": "X", "country": "CZ",
         "status": "Existing", "photo": ph} for n in (1, 2, 3, 4)]}), encoding="utf-8")
    check(call(gate_runner, "images", rate_path) == 0,
          "P1-6: a tracker/record-only all-placeholder longlist SHIPS (no brochure source)")
    # (b) but when brochures WERE examined (placeholderAudit present, even 0 candidates)
    #     a high placeholder rate is a harvest failure -> BLOCKS until reviewed.
    exam_dir = work / "neg_rate_exam"; exam_dir.mkdir()
    exam_path = exam_dir / "canonical.json"
    exam_path.write_text(json.dumps({
        "meta": {"placeholderAudit": {str(n): {"source": f"b{n}.pdf", "locator": "page 1",
                                               "candidates": 0, "files": []} for n in (1, 2, 3, 4)}},
        "properties": [
            {"id": n, "park": f"P{n}", "developer": "D", "city": "X", "country": "CZ",
             "status": "Existing", "photo": ph} for n in (1, 2, 3, 4)]}), encoding="utf-8")
    check(call(gate_runner, "images", exam_path) != 0,
          "P1-6: an all-placeholder longlist where brochures WERE examined still BLOCKS")
    (exam_dir / "placeholder_audit_ack.json").write_text(
        json.dumps({"placeholder_rate_ok": True}), encoding="utf-8")
    check(call(gate_runner, "images", exam_path) == 0,
          "the recorded G-images sign-off unblocks a reviewed placeholder rate")

    print("Stage 5/6 - build + post-build gates:")
    check(call(build_dashboard, canonical, "--out", built) == 0, "build completes")
    check(call(gate_runner, "validate-html", built, "--canonical", canonical) == 0,
          "validate-html ALL-PASS (byte-identical chrome)")
    # S6-8: on the byte-identical pass the JSON re-parse is skipped, but a byte-DIVERGED file
    # (here a corrupted PROPS block) must still BLOCK - the inequality path (and the byte-equality
    # floor) still bite, and the nested JSON loop does not crash.
    _bad_html = work / "built_badjson.html"
    _bad_html.write_text(built.read_text(encoding="utf-8").replace("const PROPS = [", "const PROPS = [{", 1),
                         encoding="utf-8")
    check(call(gate_runner, "validate-html", _bad_html, "--canonical", canonical) != 0,
          "S6-8: validate-html still BLOCKS a byte-diverged file with a corrupted data block")
    check(call(gate_runner, "reconcile", built, "--canonical", canonical) == 0,
          "reconcile ALL-PASS")
    # G-i18n deterministic floor: the fixture builds English chrome, so the floor PASSES
    # (UI complete, well-formed LOCALE, no unfilled token, placeholders intact; the
    # silent-fallback check is skipped for EN). Then prove the gate BITES: a built file
    # whose const UI dropped a chrome key must BLOCK.
    check(call(gate_runner, "i18n", built, "--canonical", canonical) == 0,
          "G-i18n floor ALL-PASS (complete EN chrome)")
    i18n_bad = work / "built_i18n_bad.html"
    _bad_html = built.read_text(encoding="utf-8").replace('"tab_grid":"Grid",', "", 1)
    i18n_bad.write_text(_bad_html, encoding="utf-8")
    check(call(gate_runner, "i18n", i18n_bad, "--canonical", canonical) != 0,
          "G-i18n floor BLOCKS a built file whose const UI dropped a chrome key")

    print("Stage 7 - deliver + final gate (verdict protocol):")
    deliverables = work / "deliverables"
    check(call(deliver, "--canonical", canonical, "--html", built, "--ledger", ledger_csv,
               "--out-dir", deliverables, "--slug", "Fixture") == 0, "deliver completes")

    # the flat Longlist workbook ships alongside the ledger: ONE row per property,
    # variables in columns, with a DERIVED monthly-rent column (xlsx; csv fallback).
    ll_xlsx, ll_csv = deliverables / "Fixture_Longlist.xlsx", deliverables / "Fixture_Longlist.csv"
    ll = ll_xlsx if ll_xlsx.exists() else ll_csv
    check(ll.exists(), "deliver writes the flat Longlist workbook (xlsx or csv fallback)")
    if ll.exists():
        if ll.suffix == ".xlsx":
            try:
                from openpyxl import load_workbook
                ws = load_workbook(ll).active
                data_rows = ws.max_row - 1
                hdr = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
            except Exception as e:
                data_rows, hdr = -1, []
                print(f"    (longlist read crash: {e})")
        else:
            import csv as _csv
            rr = list(_csv.reader(open(ll, newline="", encoding="utf-8")))
            data_rows, hdr = len(rr) - 1, (rr[0] if rr else [])
        check(data_rows == len(data["properties"]),
              f"Longlist has ONE row per property ({len(data['properties'])}, got {data_rows})")
        check(hdr[:1] == ["ID"] and "Warehouse rent (monthly)" in hdr
              and "Total annual rent" in hdr and "Total monthly rent" in hdr,
              "Longlist is property-per-row (col 1 = 'ID') with derived monthly + total rent columns")

    # v14: total rent = GLA x rate (split warehouse+office when a separate office rate exists)
    check(deliver._total_rent({"warehouseRentVal": 60.0, "warehouseArea": 40000, "rentUnit": "€/sq m/yr"}) == "€ 2,400,000 / yr",
          "v14: total annual = warehouse area x rate when there is no office split")
    check(deliver._total_rent({"warehouseRentVal": 60.0, "warehouseArea": 40000, "officeAreaVal": 5000, "rentUnit": "€/sq m/yr"}) == "€ 2,700,000 / yr",
          "v14: office area joins total GLA at the warehouse rate when no separate office rate")
    check(deliver._total_rent({"warehouseRentVal": 60.0, "warehouseArea": 40000, "officeAreaVal": 5000, "officeRentVal": 100.0, "rentUnit": "€/sq m/yr"}) == "€ 2,900,000 / yr",
          "v14: a separate office rate SPLITS the calc (40000x60 + 5000x100 = 2,900,000)")
    check(deliver._total_rent({"warehouseArea": 40000, "rentUnit": "€/sq m/yr"}) == "tbd",
          "v14: total rent is tbd without a warehouse rate")
    check(deliver._total_rent({"warehouseRentVal": 60.0, "warehouseArea": 40000, "rentUnit": "€/sq m/yr"}, monthly=True) == "€ 200,000 / mo",
          "v14: total monthly = total annual / 12")

    check(call(gate_runner, "freeze", canonical) == 0, "freeze snapshot written")

    good = work / "reviews"; good.mkdir()
    (good / "G-honesty.md").write_text("All sampled values check out.\nVERDICT: green\n", encoding="utf-8")
    (good / "G-trace.md").write_text(  # prose mentioning a red verdict must NOT block
        "If a sampled field had not matched its locator, the verdict: red would apply.\n"
        "- [LOW] property=2 field=clearHeight issue=not stated action=chase landlord\n"
        "VERDICT: amber\n", encoding="utf-8")
    (good / "G-images.md").write_text("Montage reviewed, placeholders honest.\nVERDICT: green\n", encoding="utf-8")
    (good / "G-visual.md").write_text("Grid/modal/map render cleanly.\nVERDICT: green\n", encoding="utf-8")
    rc = call(final_gate, "--canonical", canonical, "--html", built,
              "--deliverables", deliverables, "--reviews", good)
    check(rc == 0, "V1/V3: final gate passes with an amber verdict + prose 'verdict: red' mention")

    # S7-17: a stub/empty deliverables set must NOT pass 'present' (size + Longlist checks)
    stub = work / "deliverables_stub"; stub.mkdir()
    (stub / "x.html").write_text("<html></html>", encoding="utf-8")  # < 5 KB stub, no ledger/gaps/longlist
    check(call(final_gate, "--canonical", canonical, "--html", built,
               "--deliverables", stub, "--reviews", good) != 0,
          "S7-17: final gate BLOCKS a stub/empty deliverables set (size + Longlist checked)")

    bad = work / "reviews_bad"; bad.mkdir()
    for f in ("G-trace.md", "G-images.md", "G-visual.md"):
        (bad / f).write_text("VERDICT: green\n", encoding="utf-8")
    (bad / "G-honesty.md").write_text("Looked fine overall, nothing to add.\n", encoding="utf-8")  # NO verdict line
    rc = call(final_gate, "--canonical", canonical, "--html", built,
              "--deliverables", deliverables, "--reviews", bad)
    check(rc != 0, "V2: a review file with no VERDICT line fails the final gate")

    check(final_gate.parse_verdict("VERDICT: RED\n") == "red"
          and final_gate.parse_verdict("notes only") is None
          and final_gate.parse_verdict("**VERDICT:** amber") == "amber",
          "parse_verdict unit checks")

    if FAILS:
        print(f"\nFIXTURE TEST: FAIL ({len(FAILS)})")
        for f in FAILS:
            print(f"  - {f}")
        return 1
    print("\nFIXTURE TEST: PASS (pipeline output clears its own gates first try; gates still bite)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
