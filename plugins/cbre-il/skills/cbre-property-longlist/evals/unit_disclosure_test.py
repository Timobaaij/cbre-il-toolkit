#!/usr/bin/env python3
"""unit_disclosure_test.py - an ASSUMED area unit says so in the workbook. (B38)

The Longlist workbook is the artefact the broker forwards, and an "Area unit" cell reading a
bare "sq ft" presents a 10.76x risk as fact when the source stated nothing.

THIS SUITE ALSO PERMANENTLY RECORDS A REJECTED DESIGN. The tempting fix was to write the area
as a string - `"12,500 (unit not stated)"` - which would have rendered honestly with no chrome
change, because the v28 `areaStr()` passes a string through unchanged. It is unsafe on three
counts, and assertion 3 below is what catches all of them:
  1. the canonical schema is {"type": "number"}, so a string HARD-BLOCKS validate-data;
  2. the jsonschema-absent fallback type-checks it too, so there is no degraded escape;
  3. if the schema were widened, `initSizeSlider` floors the size filter at the smallest
     NUMERIC area, so a non-numeric one can never pass filterList - the property VANISHES from
     the grid, the map and Compare while still being counted in the KPIs.
It would trade a cosmetic wrong label for a missing option. The card half is TEMPLATE-only:
the chrome's AREA_UNIT is DATASET-wide, so no data-side change can move it at all.

Drives real merge.main, real deliver.longlist_xlsx and real validate-data. Offline."""
from __future__ import annotations
import json
import subprocess
import sys
import tempfile
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
HELPERS = ROOT / "helpers"
sys.path.insert(0, str(HELPERS))
import _common as C  # noqa: E402
import deliver as D  # noqa: E402


def _r(src, stype, **kw):
    r = {"city": "Corby", "country": "GB", "developer": "Prologis",
         "__meta": {"source_file": src, "source_type": stype, "locator_base": "page 1"}}
    r.update(kw)
    return r


def main() -> int:
    fails = []

    def ck(ok, l):
        print(f"  [{'PASS' if ok else 'FAIL'}] {l}")
        if not ok:
            fails.append(l)

    d = Path(tempfile.mkdtemp(prefix="cbre_ud_"))
    (d / "inputs").mkdir()
    recs = [
        # STATED unit - its cell must stay the bare unit, or the disclosure over-fires
        _r("Stated.pdf", "pdf", park="Stated Park", warehouseArea=120000, areaUnit="sq ft"),
        _r("Stated2.pdf", "pdf", park="Stated Park Two", warehouseArea=95000, areaUnit="sq ft"),
        # SILENT - a numeric area and no unit anywhere
        _r("Silent.pdf", "pdf", park="Silent Park", warehouseArea=12500),
    ]
    (d / "r.json").write_text(json.dumps(recs), encoding="utf-8")
    p = subprocess.run([sys.executable, str(HELPERS / "merge.py"), "--records", str(d / "r.json"),
                        "--source-dir", str(d / "inputs"), "--out", str(d / "c.json"),
                        "--ledger", str(d / "l.csv")],
                       capture_output=True, text=True, errors="replace")
    ck((d / "c.json").exists(), f"merge completes {ascii((p.stdout + p.stderr)[-160:])}")
    if not (d / "c.json").exists():
        print(f"\nUNIT DISCLOSURE TEST: FAIL ({len(fails)})")
        return 1
    canon = json.loads((d / "c.json").read_text(encoding="utf-8"))
    props = canon["properties"]
    ua = (canon.get("meta") or {}).get("unitAssumptions") or []

    # 1. the assumption joins by ID, never by park name
    ck(bool(ua), f"the silent record is recorded as an assumption ({len(ua)})")
    ck(all(a.get("id") is not None for a in ua), "every assumption carries an id")
    silent_ids = {a["id"] for a in ua}
    silent = [q for q in props if q.get("id") in silent_ids]
    ck(len(silent) == 1 and silent[0]["park"] == "Silent Park",
       f"the id resolves to the right property {ascii(str([q.get('park') for q in silent]))}")

    # 2. the workbook cell says 'assumed' - and does NOT over-fire on the stated ones
    D.longlist_xlsx(canon, d / "ll.xlsx")
    out = d / "ll.xlsx" if (d / "ll.xlsx").exists() else d / "ll.csv"
    ck(out.exists(), f"the workbook is written {ascii(out.name)}")
    _keys = [k for k, _ in D.LONGLIST_COLUMNS]
    i_unit, i_park = _keys.index("areaUnit"), _keys.index("park")
    text = out.read_text(encoding="utf-8", errors="replace") if out.suffix == ".csv" else ""
    if out.suffix == ".csv":
        ck("assumed - source stated none" in text,
           "the CSV fallback carries the disclosure")
        ck(text.count("assumed - source stated none") == 1,
           "exactly ONE row is disclosed - it cannot over-fire onto the stated ones")
    else:
        try:
            from openpyxl import load_workbook
            ws = load_workbook(out, read_only=True).active
            rows = list(ws.iter_rows(values_only=True))
            # column 0 is the ID; the park name is its own column
            got = {r[i_park]: r[i_unit] for r in rows[1:] if r and r[i_park]}
            ck("assumed" in str(got.get("Silent Park", "")),
               f"the SILENT row's Area unit says assumed {ascii(str(got.get('Silent Park')))}")
            ck(str(got.get("Stated Park", "")).strip() == "sq ft",
               f"a STATED row keeps the BARE unit {ascii(str(got.get('Stated Park')))}")
        except ImportError:
            print("  [SKIP] openpyxl absent - CSV path covered above")

    # 3. THE STANDING GUARD, and the assertion that rejects the string design
    for q in props:
        wa = q.get("warehouseArea")
        ck(wa is None or isinstance(wa, (int, float)) and not isinstance(wa, bool),
           f"{q['park']}: warehouseArea is NUMERIC, never a string {ascii(str(wa))}")
    ck(not [k for q in props for k in q if k.endswith(("Assumed", "Suspect"))],
       "no internal flag reached a property (B05's sweep still holds)")
    ck(C.schema_errors(canon) == [] if hasattr(C, "schema_errors") else True,
       "the canonical passes its own schema")
    g = subprocess.run([sys.executable, str(HELPERS / "gate_runner.py"), "validate-data",
                        str(d / "c.json")], capture_output=True, text=True, errors="replace")
    ck(g.returncode == 0,
       f"validate-data exits 0 - a string area would have BLOCKED here {ascii(g.stdout[-120:])}")

    # 4. the dashboard must be byte-identical: this change touches meta + a workbook cell only
    import build_dashboard as B
    h1, _ = B.render(canon)
    canon2 = json.loads((d / "c.json").read_text(encoding="utf-8"))
    canon2["meta"].pop("unitAssumptions", None)
    h2, _ = B.render(canon2)
    ck(h1 == h2,
       "the built dashboard is BYTE-IDENTICAL with and without the disclosure (meta is not "
       "injected; the card half is a TEMPLATE item because AREA_UNIT is dataset-wide)")

    if fails:
        print(f"\nUNIT DISCLOSURE TEST: FAIL ({len(fails)})")
        return 1
    print("\nUNIT DISCLOSURE TEST: PASS")
    return 0


if __name__ == "__main__":
    sys.exit(main())
