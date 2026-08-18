#!/usr/bin/env python3
"""
Stage 4 - client-facing Excel (DETERMINISTIC WRITER).

Reads properties/_dataset.json (already enriched by the model) and writes a clean,
client-ready workbook. It only FORMATS - all judgement (rent, specs, what to include)
was made upstream by the model. Column choice here is authored by the model, but the
per-cell values are a straight mapping from property.json, so it can't mis-guess data.

- Two sheets: "Longlist" (all) and "For Sale" (the for-sale subset).
- Merged header bands (Property / Terms / Specification / Agent / Media) over column headers.
- Agent-quoted rent already wins (property.json.rent); no provenance shown to the client.
- URLs render as the word "link" and point to ONLINE sources (brochure PDF, website, video,
  lead photo) so they work for a client who does not have the local media folders.
"""
import os, sys, re, argparse
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from common import load_config, read_json, dedupe_props
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# (header, key/getter, width, number_format, is_link)
def display_name(rec):
    f = rec["folder"]
    f = re.sub(r"^\d+\s*-\s*", "", f)
    f = re.sub(r"\s*-\s*[A-Z]{1,2}\d[A-Z\d]?\s*\d[A-Z]{2}\s*$", "", f)
    return rec.get("address", {}).get("name") or f

def rent_cell(rec):
    r = rec.get("rent") or {}
    if r.get("value") is not None:
        return r["value"], "£#,##0.00"
    return (r.get("text") or "On application"), None

def agent_str(rec):
    org = rec.get("agent_organisation") or ""
    ags = rec.get("agents") or []
    nm = ags[0]["name"] if ags and ags[0].get("name") else ""
    return " – ".join([x for x in [org, nm] if x])

def contact_str(rec):
    ags = rec.get("agents") or []
    if not ags:
        return ""
    a = ags[0]
    tel = a.get("mobile") or a.get("tel") or ""
    return " / ".join([x for x in [tel, a.get("email")] if x])

# Column groups: (band, [(header, fn, width, numfmt, link_kind)])
def street(rec):
    a = rec.get("address") or {}
    return ", ".join([x for x in [a.get("line1"), a.get("line2"), a.get("town")] if x])

GROUPS = [
    ("Property", [
        ("#", lambda r: r["order"], 4, "0", None),
        ("Property", display_name, 28, None, None),
        ("Address", street, 34, None, None),
        ("Postcode", lambda r: r.get("postcode"), 10, None, None),
        ("Area", lambda r: r.get("area") or (r.get("address") or {}).get("town"), 20, None, None),
        ("Latitude", lambda r: ((r.get("coordinates") or {}).get("map") or {}).get("lat"), 11, "0.000000", None),
        ("Longitude", lambda r: ((r.get("coordinates") or {}).get("map") or {}).get("lng"), 11, "0.000000", None),
    ]),
    ("Terms", [
        ("Tenure", lambda r: r.get("tenure"), 15, None, None),
        # GLA is the total the agent quotes; warehouse and office are its two parts, so
        # warehouse + office returns the GLA. Warehouse falls back to the GLA and office stays
        # blank where no source states the split (never guessed).
        ("Total GLA (sq ft)", lambda r: (r.get("size") or {}).get("sqft"), 13, "#,##0", None),
        ("Warehouse (sq ft)", lambda r: ((r.get("size") or {}).get("warehouse_sqft")
                                         or (r.get("size") or {}).get("sqft")), 14, "#,##0", None),
        ("Office (sq ft)", lambda r: (r.get("size") or {}).get("office_sqft"), 12, "#,##0", None),
        ("Availability", lambda r: (r.get("spec") or {}).get("availability"), 20, None, None),
        ("Rent (£/sq ft)", rent_cell, 16, None, None),
        ("Basis", lambda r: (r.get("rent") or {}).get("basis"), 20, None, None),
    ]),
    ("Specification", [
        ("Clear height", lambda r: (r.get("spec") or {}).get("clear_height"), 12, None, None),
        ("Power", lambda r: (r.get("spec") or {}).get("power"), 15, None, None),
        ("Loading", lambda r: (r.get("spec") or {}).get("loading"), 20, None, None),
        ("Yard", lambda r: (r.get("spec") or {}).get("yard"), 14, None, None),
        ("Parking", lambda r: (r.get("spec") or {}).get("parking"), 13, None, None),
        ("EPC", lambda r: (r.get("spec") or {}).get("epc"), 9, None, None),
        ("BREEAM", lambda r: (r.get("spec") or {}).get("breeam"), 13, None, None),
    ]),
    ("Overview", [
        ("Summary", lambda r: r.get("summary"), 55, None, None),
    ]),
    ("Media (online)", [
        ("Brochure", lambda r: (r.get("media") or {}).get("brochure_url"), 9, None, "url"),
        ("Website", lambda r: (r.get("links") or {}).get("website"), 9, None, "url"),
        ("Video", lambda r: (r.get("media") or {}).get("video_url"), 8, None, "url"),
        ("Photo", lambda r: (r.get("media") or {}).get("lead_image_url"), 8, None, "url"),
    ]),
]

ARIAL = Font(name="Arial", size=10)
HFONT = Font(name="Arial", size=10, bold=True, color="FFFFFF")
BAND = Font(name="Arial", size=10, bold=True, color="FFFFFF")
HFILL = PatternFill("solid", fgColor="17472A")
BANDFILL = PatternFill("solid", fgColor="0E3320")
LINKFONT = Font(name="Arial", size=10, color="0563C1", underline="single")
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = {"Loading", "Availability", "Agent", "Contact", "Area", "Property", "Basis"}

def write_sheet(ws, records, props_root, overrides=None):
    """`overrides` is {property name: {column header: literal value}} - a client-edited sheet
    treated as ground truth for those cells. Used for values that can only be TEXT in the
    workbook ("Up to 315,000", "TBC") while the dashboard needs a real number in the same field."""
    overrides = overrides or {}
    cols = [(band, h, fn, w, nf, lk) for band, items in GROUPS for (h, fn, w, nf, lk) in items]
    # row 1 = bands (merged), row 2 = headers
    ci = 1
    for band, items in GROUPS:
        span = len(items)
        ws.merge_cells(start_row=1, start_column=ci, end_row=1, end_column=ci + span - 1)
        c = ws.cell(1, ci, band); c.font = BAND; c.fill = BANDFILL
        c.alignment = Alignment(horizontal="center", vertical="center")
        ci += span
    for idx, (band, h, fn, w, nf, lk) in enumerate(cols, 1):
        c = ws.cell(2, idx, h); c.font = HFONT; c.fill = HFILL; c.border = BORDER
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(idx)].width = w
    for ri, rec in enumerate(records, 3):
        folder = rec["folder"]
        for idx, (band, h, fn, w, nf, lk) in enumerate(cols, 1):
            cell = ws.cell(ri, idx); cell.border = BORDER
            if lk == "url":
                target = fn(rec)
                if target:
                    cell.value = "link"; cell.hyperlink = target; cell.font = LINKFONT
                    cell.alignment = Alignment(horizontal="center", vertical="top")
                continue
            val = fn(rec)
            numfmt = nf
            if h == "Rent (£/sq ft)":
                val, numfmt = val  # rent_cell returns (value, fmt)
            ov = overrides.get(display_name(rec), {})
            if h in ov:                      # client-edited cell wins, verbatim
                val, numfmt = ov[h], None
            cell.value = val
            cell.font = ARIAL
            cell.alignment = Alignment(vertical="top", wrap_text=(h in WRAP),
                                       horizontal=("right" if numfmt else "left"))
            if val is not None and numfmt:
                cell.number_format = numfmt
    ws.freeze_panes = "C3"
    ws.auto_filter.ref = f"A2:{get_column_letter(len(cols))}2"
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 18
    ws.row_dimensions[2].height = 30

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--config", required=True)
    ap.add_argument("--out", default=None)
    args = ap.parse_args()
    cfg = load_config(args.config)
    work = cfg["work_dir"]
    props_root = os.path.join(work, "properties")
    ds = read_json(os.path.join(props_root, "_dataset.json"), {}) or {}
    # dedupe: the same physical unit often arrives as several Kato match-request rows
    # (one per broker who quoted it) - the client Excel must list each opportunity once,
    # consistent with the toolkit dashboard, which applies the identical grouping.
    recs = dedupe_props(ds.get("properties", []))
    out = args.out or os.path.join(work, "Kato Longlist (Client).xlsx")
    # optional client-edited display values, treated as ground truth for the cells they cover
    overrides = read_json(os.path.join(work, "excel_display_overrides.json"), {}) or {}

    wb = Workbook()
    ws1 = wb.active; ws1.title = "Longlist"
    write_sheet(ws1, recs, props_root, overrides)
    sales = [r for r in recs if r.get("for_sale")]
    if sales:
        write_sheet(wb.create_sheet("For Sale"), sales, props_root, overrides)
    wb.save(out)
    print(f"Saved {out} | Longlist rows={len(recs)} | For Sale rows={len(sales)}", flush=True)

if __name__ == "__main__":
    main()
